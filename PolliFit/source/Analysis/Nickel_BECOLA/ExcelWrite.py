'''
Created on 06/06/2019

@author: fsommer
'''
from openpyxl import Workbook, load_workbook  # Needed to read/write data to xlsx
from string import ascii_uppercase
import os

import csv, os, sqlite3, re

import numpy as np

from pylab import plot, meshgrid, cm, imshow, contour, clabel, colorbar, axis, title, show


class ExcelWriter():
    '''
    This object writes information to an excel file which either already exists or will be created
    '''

    def __init__(self, xlsx_filepath):
        super(ExcelWriter, self).__init__()

        self.file = xlsx_filepath

        # load or create workbook
        try:
            # try to load existing workbook
            self.wb = self.load_single_xlsx_file(self.file)
        except:
            # if no xlsx file exists yet, create empty workbook
            self.wb = Workbook()

        # create variable for active worksheet
        self.active_sheet = self.wb.active

        # create variable to store last used row
        self.last_row = 9  # starts with 9 from template

    def load_single_xlsx_file(self, filepath):
        # print('Loading .xlsx file from %s' % str(filepath))
        new_wb = load_workbook(filepath, data_only=True)  # data only removes formulas
        return new_wb

    def create_new_worksheet(self, ws_name):
        """
        creates a new worksheet and inserts at the end
        :param ws_name: str: name of new worksheet
        :return:
        """
        self.wb.create_sheet(ws_name)