from Measurement.XMLImporter import XMLImporter
from openpyxl import Workbook
import Tools
import os
import ast
import matplotlib.pyplot as plt
import numpy as np
import sqlite3
import BatchFit
import Physics
from scipy import constants as const
from scipy.optimize import curve_fit
from operator import add


class NiAnalysis:
    def __init__(self,calT, calG, systU):
        self.lastRow = 1
        self.workingdir = 'D:\\Daten\\IKP\\Nickel-Auswertung\\Auswertung'
        self.filepath = os.path.join(self.workingdir, 'data')
        self.db = os.path.join(self.workingdir, 'Nickel_BECOLA.sqlite')
        print('Database:', self.db)
        #self.runs = ['AsymVoigt0', 'AsymVoigt1', 'AsymVoigt2', 'AsymVoigtAll']
        self.runs = ['AsymVoigt0', 'AsymVoigt1', 'AsymVoigt2']
        self.run = ''
        self.linVars = ['58_0','58_1','58_2','58_All']
        self.run = ''
        #self.files = ['Scaler0.xlsx', 'Scaler1.xlsx','Scaler2.xlsx','AllScalers.xlsx']
        self.files = ['Scaler0.xlsx', 'Scaler1.xlsx', 'Scaler2.xlsx']
        self.xlsFile = ''
        self.calTuples = calT
        self.calGroups56 = calG
        self.shifts56 = []
        self.uncert56 = []
        self.systUncert = systU
        self.allRes = []
        self.allUnc = []
        self.allMeas = []
        self.spec = XMLImporter(path=self.workingdir + '\\data\\' + 'BECOLA_6501.xml')

    def createWB(self):
        self.wb = Workbook()
        self.wsCal = self.wb.create_sheet('Calibration', 0)
        self.wsCal.title = 'Calibration'
        self.wsCal.cell(row=self.lastRow, column=1, value='Tuple')
        self.wsCal.cell(row=self.lastRow, column=2, value='old Voltage')
        self.wsCal.cell(row=self.lastRow, column=3, value='old isotope shift')
        self.wsCal.cell(row=self.lastRow, column=4, value='calibrated Voltage')
        self.wsCal.cell(row=self.lastRow, column=5, value='calibrated istope shift')
        self.wsIS = self.wb.create_sheet('Isotope shifts 56Ni', 1)
        self.wsIS.title = 'Isotope shifts 56Ni'
        self.wsIS.cell(row=1, column=1, value='File')
        self.wsIS.cell(row=1, column=2, value='Isotope shift')
        self.wsIS.cell(row=1, column=3, value='Uncertainty')
        self.wb.save(self.xlsFile)

    def plotSpec(self, filelist):
        for file in filelist:
            spec = XMLImporter(path=os.path.join(self.filepath, file))
            Tools.add_missing_columns(self.db)
            voltage = spec.x[0]
            ctsScaler0 = spec.cts[0][0]
            ctsScaler1 = spec.cts[0][1]
            ctsScaler2 = spec.cts[0][2]

            plt.plot(voltage, ctsScaler0)
            plt.show()

    def calIShift(self,fileTuple):
        # Get center of reference
        file58 = 'BECOLA_' + str(fileTuple[0]) + '.xml'
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        # Get corresponding isotope
        cur.execute('''SELECT type FROM Files WHERE file = ? ''', (file58,))
        iso_type58 = cur.fetchall()[0][0]
        # Query fitresults for file and isotope combo
        cur.execute('''SELECT pars FROM FitRes WHERE file = ? AND iso = ? AND run = ?''', (file58, iso_type58, self.run))
        pars58 = cur.fetchall()
        con.close()
        pars58dict = ast.literal_eval(pars58[0][0])
        center58 = pars58dict['center']

        # Get center of interested isotope
        file60 = 'BECOLA_' + str(fileTuple[1]) + '.xml'
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        # Get corresponding isotope
        print(file60)
        cur.execute(
            '''SELECT type FROM Files WHERE file = ? ''', (file60,))
        iso_type60 = cur.fetchall()[0][0]
        # Query fitresults for file and isotope combo
        cur.execute('''SELECT pars FROM FitRes WHERE file = ? AND iso = ? AND run = ?''', (file60, iso_type60, self.run))
        pars60 = cur.fetchall()
        con.close()
        pars60dict = ast.literal_eval(pars60[0][0])
        center60 = pars60dict['center']

        # Calculate isotope shift
        ishift = center60[0] - center58[0]
        print('Isotope shift is', ishift)

        # Calculate uncertainty
        uncert = np.sqrt(np.square(center58[1]) + np.square(center60[1]))

        return ishift, uncert

    def assignAsy(self, tup):
        file58 = 'BECOLA_' + str(tup[0]) + '.xml'
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT pars from FitRes WHERE file = ?''', (file58,))
        pars = cur.fetchall()
        pars = ast.literal_eval(pars[0][0])
        asy = pars['asy'][0]
        con.close()
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Lines SET shape = ?''', (
        "{'name': 'VoigtAsy', 'sigma': 200, 'offset': 343.0, 'asy': " + str(
            asy) + " , 'gamma': 20, 'offsetSlope': 0}",))
        con.close()

    def singleCaliOnIS(self):
        files = []
        shifts = []
        unc = []
        for tup in self.calTuples:
            files.append(tup[0])
            print('Calibrating files', tup)
            ishift, err = self.calIShift(tup)
            shifts.append(ishift)
            unc.append(err)

            # Get acceleration voltage
            file58 = 'BECOLA_' + str(tup[0]) + '.xml'
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (file58,))
            accVolt = cur.fetchall()[0][0]
            con.close()
            print('Current acceleration voltage is', accVolt)

            # Open excel
            # load or create workbook
            self.lastRow = self.wsCal.max_row

            self.wsCal.cell(row=self.lastRow+1, column=1, value=str(tup))
            self.wsCal.cell(row=self.lastRow+1, column=2, value=accVolt)
            self.wsCal.cell(row=self.lastRow+1, column=3, value=ishift)

            # Calculate differential Doppler shift
            diffDopp58 = Physics.diffDoppler(850343799, accVolt, 58)
            diffDopp60 = Physics.diffDoppler(850343799, accVolt, 60)

            # Calculate calibration
            calVolt = accVolt + (ishift - 509.1)/(diffDopp60-diffDopp58)
            print('Calibrated acceleration voltage is', calVolt)

            # Update acceleration voltage
            file60 = 'BECOLA_' + str(tup[1]) + '.xml'
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''UPDATE Files SET accVolt = ? WHERE file LIKE ? OR file LIKE ?''', (calVolt,file58,file60))
            con.commit()
            con.close()

            self.wsCal.cell(row=self.lastRow+1, column=4, value=calVolt)
            self.wb.save(self.xlsFile)

        return shifts,files, unc

    def singleCaliOnAbs(self):
        files = []
        abs_frequ = []
        unc = []
        files58 = self.getFiles('58Ni')
        ref_frequ = 0
        for lv in self.linVars:
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT frequency FROM Lines WHERE lineVar = ? ''', (lv,))
            ref_frequ += cur.fetchall()[0][0]
            con.close()
        ref_frequ = ref_frequ / len(self.linVars)

        for f in files58:
            files.append(int(f[7:11]))
            print('Calibrating files', f[7:11])
            center_frequ = []
            weights = []
            # Get center
            for run in self.runs:
                con = sqlite3.connect(self.db)
                cur = con.cursor()
                cur.execute('''SELECT pars FROM FitRes WHERE file = ? AND run = ?''', (f, run,))
                pars = cur.fetchall()[0][0]
                con.close()
                center_frequ.append(ast.literal_eval(pars)['center'][0])
                weights.append(1 / (ast.literal_eval(pars)['center'][1] ** 2))
            center_mean = np.average(center_frequ, weights=weights)
            center = ref_frequ + center_mean
            abs_frequ.append(ref_frequ + center_mean)
            unc.append(np.std(center_frequ))
            print('Absolute center frequency is:', center)

            # Get acceleration voltage
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (f,))
            accVolt = cur.fetchall()[0][0]
            con.close()
            print('Current acceleration voltage is', accVolt)

            # Open excel
            # load or create workbook
            self.lastRow = self.wsCal.max_row
            self.wsCal.cell(row=self.lastRow+1, column=1, value=str(f))
            self.wsCal.cell(row=self.lastRow+1, column=2, value=accVolt)
            self.wsCal.cell(row=self.lastRow+1, column=3, value=center)

            # Calculate differential Doppler shift
            diffDopp58 = Physics.diffDoppler(center, accVolt, 58)

            # Calculate calibration
            calVolt = accVolt + (center - 850343678) / diffDopp58
            print('Calibrated acceleration voltage is', calVolt)

            self.wsCal.cell(row=self.lastRow+1, column=4, value=calVolt)
            self.wb.save(self.xlsFile)
        print(abs_frequ, files, unc)
        return abs_frequ,files, unc

    def calibrateOnIS(self):

        # reset
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Files SET accVolt = ?''', (29850,))
        con.commit()
        con.close()

        # Adjust center in isotopes
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (-200, '58Ni',))
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (300, '60Ni',))
        con.commit()
        con.close()

        filelist58 = self.getFiles('58Ni')
        print('Files to fit:',filelist58)
        self.fitAll(filelist58)

        filelist60 = self.getFiles('60Ni')
        self.fitAll(filelist60)

        shifts, files, uncert = self.singleCaliOnIS()

        plt.errorbar(files, shifts, yerr=uncert,fmt = 'bo')
        #plt.plot(files,shifts,'bo')
        plt.title('Uncalibrated Reference Shift')
        plt.show()
        print('Uncalibrated Reference shifts:', shifts)

        #Adjust center in isotopes
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (-400, '58Ni',))
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (100, '60Ni',))
        con.commit()
        con.close()

        self.fitAll(filelist58)
        self.fitAll(filelist60)
        shifts = []
        uncert = []
        for tup in self.calTuples:
            print('Calculating calibrated isotope shift:')
            ishift, err = self.calIShift(tup)
            print('Calibrated isotope shift of File', tup[0], 'is', ishift)
            shifts.append(ishift)
            uncert.append(err)
        plt.errorbar(files, shifts, yerr=uncert, fmt='bo')
        #plt.plot(files, shifts, 'bo')
        plt.title('Calibrated Reference Shift')
        plt.show()

        # Save calibrated istotope shifts to excel
        r = 2
        for s in shifts:
            self.wsCal.cell(row=r, column=5, value=s)
            r += 1
        self.wb.save(self.xlsFile)

    def calibrateOnAbs(self):

        # reset
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Files SET accVolt = ?''', (29850,))
        con.commit()
        con.close()

        # Adjust center in isotopes
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (-200, '58Ni',))
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (300, '60Ni',))
        con.commit()
        con.close()

        filelist58 = self.getFiles('58Ni')
        print('Files to fit:',filelist58)
        self.fitAll(filelist58)

        filelist60 = self.getFiles('60Ni')
        self.fitAll(filelist60)

        center, files, uncert = self.singleCaliOnAbs()

        #plt.errorbar(files, center, yerr=uncert,fmt = 'bo')
        plt.plot(files,center,'bo')
        plt.title('Uncalibrated Reference transition')
        plt.show()
        print('Uncalibrated Reference absolute transition frequency:', center)

        #Adjust center in isotopes
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (-400, '58Ni',))
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (100, '60Ni',))
        con.commit()
        con.close()

        for tup in self.calTuples:
            file58 = 'BECOLA_' + str(tup[0]) + '.xml'
            file60 = 'BECOLA_' + str(tup[1]) + '.xml'
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (file58,))
            accVolt = cur.fetchall()[0][0]
            cur.execute('''UPDATE Files SET accVolt = ? WHERE file = ?''', (accVolt, file60))
            con.commit()
            con.close()

        self.fitAll(filelist58)
        self.fitAll(filelist60)
        shifts = []
        uncert = []
        files = []
        for tup in self.calTuples:
            print('Calculating calibrated isotope shift:')
            ishift, err = self.calIShift(tup)
            print('Calibrated isotope shift of File', tup[0], 'is', ishift)
            shifts.append(ishift)
            uncert.append(err)
            files.append(tup[0])
        plt.errorbar(files, shifts, yerr=uncert, fmt='bo')
        #plt.plot(files, shifts, 'bo')
        plt.title('Calibrated Reference Shift')
        plt.show()

        # Save calibrated istotope shifts to excel
        r = 2
        for s in shifts:
            self.wsCal.cell(row=r, column=5, value=s)
            r += 1
        self.wb.save(self.xlsFile)

    def adjOffset(self, offset):
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT shape FROM Lines WHERE lineVar LIKE ? ''', (self.lineVar,))
        shape = cur.fetchall()[0][0]
        con.close()
        shapeDict = ast.literal_eval(shape)
        shapeDict['offset'] = offset
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Lines SET shape = ? ''', (str(shapeDict),))
        con.commit()
        con.close()
        print('Adjusted Offset to ', shapeDict['offset'])

    def adjCenter(self, center):
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ?''', (center,))
        con.commit()
        con.close()
        print('Adjusted Center to', center)

    def fitAll(self, files):
        for f in files:
            print('Fitting File', f)
            for tup in self.calTuples:
                if 'BECOLA_' + str(tup[1]) + '.xml' == f:
                    file58 = 'BECOLA_' + str(tup[0]) + '.xml'
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''SELECT pars from FitRes WHERE file = ? AND run = ?''', (file58, self.run))
                    pars = cur.fetchall()
                    pars = ast.literal_eval(pars[0][0])
                    asy = pars['asy'][0]
                    print('The Asymmetry Factor of file', file58 , ' and run', self.run, 'is', asy)
                    con.close()
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET shape = ?''', (
                    "{'name': 'VoigtAsy', 'sigma': 200, 'offset': 343.0, 'asy': " + str(
                        asy) + " , 'gamma': 20, 'offsetSlope': 0}",))
                    con.commit()
                    con.close()
                    con  = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET fixShape = ?''',(
                        "{'asy': True, 'offsetSlope': True, 'sigma': [0,200], 'offset': False, 'gamma': [0,50]}",))
                    con.commit()
                    con.close()
            for tup in self.calGroups56:
                if 'BECOLA_' + str(tup[1]) + '.xml' == f:
                    file58 = 'BECOLA_' + str(tup[0]) + '.xml'
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''SELECT pars from FitRes WHERE file = ? AND run = ?''', (file58, self.run))
                    pars = cur.fetchall()
                    pars = ast.literal_eval(pars[0][0])
                    asy = pars['asy'][0]
                    con.close()
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET shape = ?''', (
                    "{'name': 'VoigtAsy', 'sigma': 200, 'offset': 343.0, 'asy': " + str(
                        asy) + " , 'gamma': 20, 'offsetSlope': 0}",))
                    con.commit()
                    con.close()
                    con  = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET fixShape = ?''',(
                        "{'asy': True, 'offsetSlope': True, 'sigma': [0,200], 'offset': False, 'gamma': [0,50]}",))
                    con.commit()
                    con.close()
                    print('The Asymmetry Factor of file', file58, ' and run', self.run, 'is', asy)
            spec = XMLImporter(path=os.path.join(self.filepath, f))
            offset = (spec.cts[0][0][0] + spec.cts[0][0][-1]) / 2
            self.adjOffset(offset)
            index = list(spec.cts[0][0]).index(max(spec.cts[0][0]))
            print(f)
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (f,))
            accV = cur.fetchall()[0][0]
            con.close()
            centerV = accV - spec.x[0][index]
            print('AccV:', accV, 'DAC:', spec.x[0][index], 'CenterV:', centerV)
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT laserFreq FROM Files WHERE file = ? ''', (f,))
            laserFrequ = cur.fetchall()[0][0]
            con.close()
            print('LaserFreq:', laserFrequ)
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT type FROM Files WHERE file = ?''', (f,))
            iso = cur.fetchall()[0][0]
            con.close()
            print('Iso:', iso)
            if iso == '58Ni':
                mass = 58
            elif iso == '60Ni':
                mass = 60
            else:
                mass = 56
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT frequency FROM Lines WHERE lineVar = ? ''',('58_0',))
            frequ = cur.fetchall()[0][0]
            con.close()
            v = Physics.relVelocity(Physics.qe * centerV, mass * Physics.u)
            v = -v
            print('relDoppler:', Physics.relDoppler(laserFrequ, v))
            centerFrequ = Physics.relDoppler(laserFrequ, v) - frequ
            print('Dopplershifted Frequ:', centerFrequ)
            center = centerFrequ - 500
            self.adjCenter(center)
            print('Run to fit with:',self.run)
            BatchFit.batchFit(np.array([f]), self.db, self.run)
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''UPDATE Lines SET fixShape = ?''', (
                "{'asy': [0,20], 'offsetSlope': True, 'sigma': [0,200], 'offset': False, 'gamma': [0,50]}",))
            con.commit()
            con.close()

    def voltToFreq(self, voltage, laserFreq, mass):
        beta = np.sqrt(1 - np.square(mass * const.c**2 / (const.e * voltage + mass * const.c**2)))
        return laserFreq / np.sqrt(1 - beta**2) * (1 - beta)

    def getFiles(self, isotope):
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT file FROM Files WHERE type LIKE ? ''', (isotope,))
        files = cur.fetchall()
        con.close()
        for f in files:
            keep = False
            #print('File', f)
            for check in self.calGroups56:
                #print('Tupel', check)
                if 'BECOLA_' + str(check[1]) + '.xml' == f[0]:
                    keep = True
            #print(keep)
            #if not keep:
                #con = sqlite3.connect(self.db)
                #cur = con.cursor()
                #cur.execute('''DELETE FROM Files WHERE file = ?''', (f[0],))
                #con.commit()
                #con.close()


        return [f[0] for f in files]

    def assignCal(self):
        for tup in self.calGroups56:
            file56 = 'BECOLA_' + str(tup[1]) + '.xml'
            file58 = 'BECOLA_' + str(tup[0]) + '.xml'

            print('Getting calibrated Voltage from file', file58)

            # Query calibrated voltage
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM files WHERE file LIKE ?''',(file58,))
            calVolt = cur.fetchall()[0][0]
            con.close()

            # Update acceleration voltage to calibration
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''UPDATE Files SET accVolt = ? WHERE file LIKE ?''',(calVolt,file56,))
            con.commit()
            con.close()

            print('Voltage updated for file', file56)

    def calc56Ni(self):
        shifts = []
        uncert = []
        files = []
        r = 2
        for tup in self.calGroups56:
            print('Calculating isotope shift:')
            ishift, err = self.calIShift(tup)
            print('Isotope shift of File', tup[1], 'is', ishift)
            shifts.append(ishift)
            uncert.append(err)
            files.append(tup[1])
            self.wsIS.cell(row=r, column=1, value=str(tup[1]))
            self.wsIS.cell(row=r, column=2, value=ishift)
            self.wsIS.cell(row=r, column=3, value=err)
            r += 1
        self.wb.save(self.xlsFile)
        print(uncert)
        plt.errorbar(files, shifts, yerr=uncert, fmt='bo')
        #plt.plot(files, shifts, 'bo')
        plt.title('Isotope Shift 56Ni')
        plt.show()
        self.shifts56 = shifts
        self.uncert56 = uncert
        self.allRes.extend(shifts)
        self.allUnc.extend(uncert)

    def average(self):
        wts = []
        for i in self.uncert56:
            item = 1 / (i ** 2)
            wts.append(item)
        waverage = np.average(self.shifts56, weights=wts)
        print('Weighted average: ' + str(waverage))

        lastRow = self.wsIS.max_row
        self.wsIS.cell(row=lastRow + 2, column= 1, value='Weighted Mean')
        self.wsIS.cell(row=lastRow +2, column=2, value=waverage)

        sigma = np.std(self.shifts56)
        print('Standard deviation: ' + str(sigma))

        self.wsIS.cell(row=lastRow + 2, column=3, value=sigma)
        self.wb.save(self.xlsFile)
        n = len(self.shifts56)

        standardfehler = sigma / (n ** (1 / 2))
        print('Standard error of the mean: ' + str(standardfehler))

        werror = 1 / np.sqrt((sum(wts)))
        print('Weighted error:', werror)

        measurements = list(range(1,n + 1))

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - standardfehler, waverage + standardfehler, alpha=0.2, linewidth=0,
                         color='g')
        plt.fill_between([0, n + 1], waverage - standardfehler - self.systUncert, waverage + standardfehler + self.systUncert,
                         alpha=0.1, linewidth=0, color='r')
        plt.title('Standard error')
        plt.show()

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - sigma, waverage + sigma, alpha=0.2, linewidth=0, color='b')
        plt.fill_between([0, n + 1], waverage - sigma - self.systUncert, waverage + sigma + self.systUncert, alpha=0.2, linewidth=0,
                         color='r')
        plt.title('Standard deviation')
        plt.show()

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - werror, waverage + werror, alpha=0.2, linewidth=0, color='y')
        plt.fill_between([0, n + 1], waverage - werror - self.systUncert, waverage + werror + self.systUncert, alpha=0.2, linewidth=0,
                         color='r')
        plt.title('Weighted error')
        plt.show()

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - sigma / np.sqrt(n - 1), waverage + sigma / np.sqrt(n - 1), alpha=0.2,
                         linewidth=0, color='b')
        plt.fill_between([0, n + 1], waverage - sigma / np.sqrt(n - 1) - self.systUncert,
                         waverage + sigma / np.sqrt(n - 1) + self.systUncert, alpha=0.2, linewidth=0, color='r')
        plt.title('Standard deviation / sqrt(N-1)')
        plt.show()

    def stackFilesTProj(self, fileliste):
        sum_t_proj = np.zeros(1024)
        for file in fileliste:
            t_proj = self.findTG(file)
            sum_t_proj = list(map(add, sum_t_proj, t_proj))
        time = list(range(0, len(sum_t_proj)))
        plt.plot(time, sum_t_proj, 'b.')
        plt.title('ALL')
        a, sigma, center, offset = self.fitTime(time, sum_t_proj)
        plt.plot(time, self.gauss(time, a, sigma, center, offset), 'r-')
        plt.axvline(center - 4 * sigma, color='y')
        plt.axvline(center + 4 * sigma, color='y')
        plt.show()
        return center, sigma

    def fitTime(self, time, cts):
        start_par = np.array([15000,10,600,10700])
        param_bounds = ([0, 0, -np.inf,-np.inf], [np.inf,np.inf, np.inf, np.inf])
        a, sigma, center, offset = curve_fit(self.gauss, time, cts, start_par, bounds=param_bounds)[0]
        print(a, sigma, center, offset)
        return a, sigma, center, offset

    def stackFiles(self, filelist):
        c, s = self.stackFilesTProj(filelist)
        t_min = (c - 2 * s) / 100
        t_max = (c + 2 * s) / 100
        bin = 3
        voltage = np.arange(-261, -42, bin)
        sumcts = np.zeros(len(voltage))
        sumbg = np.zeros(len(voltage))
        scans = np.zeros(len(voltage))
        for f in filelist:
            spec = XMLImporter(path=self.workingdir + '\\data\\' + str(f),
                           softw_gates=[[-350, 0, t_min, t_max], [-350, 0, t_min, t_max], [-350, 0, t_min, t_max]])
            off = 200
            bg = XMLImporter(path=self.workingdir + '\\data\\' + str(f),
                           softw_gates=[[-350, 0, t_min +off, t_max+off], [-350, 0, t_min+off, t_max+off], [-350, 0, t_min+off, t_max+off]])
            print('error', spec.err)
            for j, x in enumerate(spec.x[0]):
                for i, v in enumerate(voltage):
                    if v - bin / 2 < x <= v + bin / 2:
                        sumcts[i] += spec.cts[0][0][j]
                        sumbg[i] += bg.cts[0][0][j]
                        scans[i] += spec.nrScans[0]
            plt.plot(spec.x[0], spec.cts[0][0])
            plt.plot(bg.x[0], bg.cts[0][0])
            #plt.hist(spec.cts[0][0], spec.x[0], histtype='step')
        plt.show()
        zeroInd = np.where(sumbg == 0)
        sumcts = np.delete(sumcts, zeroInd)
        sumbg = np.delete(sumbg, zeroInd)
        voltage = np.delete(voltage, zeroInd)
        scans = np.delete(scans, zeroInd)
        plt.plot(voltage, sumcts, 'r.')
        plt.show()
        plt.plot(voltage, sumcts / scans, 'b.')
        plt.show()
        errors = []
        for c in sumcts:
            errors.append(1/np.sqrt(c))
        print(scans)

    def gauss(self, t, a, s, t0, o):
        # prams:    t: time
        #           a: cts
        #           s: sigma
        #           t0: mid of time
        #           o: offset
        return o + a / np.sqrt(2 * np.pi * s ** 2) * np.exp(-1 / 2 * np.square((t - t0) / s))

    def getAsym(self, filelist):
        print(filelist,':')
        asy = []
        for f in filelist:
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT pars from FitRes WHERE file = ?''', (f,))
            pars = cur.fetchall()
            pars = ast.literal_eval(pars[0][0])
            print(pars['asy'][0])
            asy.append(pars['asy'][0])
            con.close()
        for a in asy:
            if a > 50:
                asy.remove(a)
        avAsy = np.average(asy)
        print(avAsy)
        plt.plot(list(range(0, len(asy))), asy, 'bo')
        plt.plot([0,40],[avAsy,avAsy])
        plt.show()

    def findTG(self,file):
        self.spec = XMLImporter(path=self.workingdir + '\\data\\' + file)
        tProj = self.spec.t_proj
        return tProj[0][0]
        # TODO alle Zeitprojketionen aufaddieren

    def calcR(self, ishift, is_uncert):
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT mass FROM Isotopes WHERE iso = ?''', ('58Ni',))
        m58 = cur.fetchall()[0][0]
        cur.execute('''SELECT mass_d FROM Isotopes WHERE iso = ?''', ('58Ni',))
        d_m58 = cur.fetchall()[0][0]
        cur.execute('''SELECT mass FROM Isotopes WHERE iso = ?''', ('56Ni',))
        m56 = cur.fetchall()[0][0]
        cur.execute('''SELECT mass_d FROM Isotopes WHERE iso = ?''', ('56Ni',))
        d_m56 = cur.fetchall()[0][0]
        con.close()

        print('Mass 58:', m58)
        print('Mass 56:', m56)

        mod_shift = ishift * m58 * m56 / (m56 - m58)

        print('Modified isotope shift:', mod_shift)

        # alpha = 0:
        k = 1270449.25353
        f = -841.096219681
        mod_dr = (mod_shift - k) / f
        dr = mod_dr * (m56 - m58) / m56 / m58
        print('dr with alpha = 0:', dr)


        # alpha = 351
        m = (m56 - m58) / m56 / m58
        print('MassModifier:', m)
        k = 975224.480423
        f = -841.096219681
        alpha = 351
        dr = (ishift - m * k) / f + m * alpha
        print('radius difference with alpha = 351', dr)
        m_uncert = d_m56 / m56 + d_m58 / m58
        k_uncert = 1518.58102368
        f_uncert = 61.0188697537
        r_uncert = np.sqrt((is_uncert / f) ** 2 +(m_uncert * (alpha - k / f)) ** 2 + (m / f * k_uncert) ** 2 +
                           (f_uncert * (ishift - m * k) / (f ** 2)) ** 2)
        print('Radius uncertainty:', r_uncert)

def Ana56Ni():
    # Calibrate Voltage
    calTuples = [(6191,6192),(6207,6208),(6242,6243),(6253,6254)]

    calGroups56 = [(6191,6202),(6191,6203),(6191,6204),(6207,6211),(6207,6213),(6207,6214),(6242,6238),(6242,6239),
                 (6242,6240),(6253,6251),(6253,6252)]
    systUncert = 2.2

    #calGroups56 = [(6242,6238),(6242,6239),(6242,6240),(6253,6251),(6253,6252)]
    #systUncert = 0.4

    #run = 'AsymVoigt0'
    #line = '58_0'
    #file = 'Scaler0.xlsx'
    #run = 'AsymVoigt1'
    #line = '58_1'
    #file = 'Scaler1.xlsx'
    #run = 'AsymVoigt2'
    #line = '58_2'
    #file = 'Scaler2.xlsx'
    #run = 'AsymVoigtAll'
    #line = '58_All'
    #file = 'AllScalers.xlsx'

    niAna = NiAnalysis(calTuples, calGroups56, systUncert)
    for i,run in enumerate(niAna.runs):
        niAna.run = run
        niAna.lineVar = niAna.linVars[i]
        niAna.xlsFile = niAna.workingdir + '\\' + niAna.files[i]
        niAna.createWB()
        niAna.calibrateOnAbs()
        niAna.assignCal()

        # Fit Ni 56
        #filelist56 = niAna.getFiles('56Ni')
        #niAna.fitAll(filelist56)
        #niAna.calc56Ni()
        #niAna.average()

    #meas = list(range(0, 20))
    #meas = list(range(0, 27))
    #print('meas:', len(meas))
    #print('Res:', len(niAna.allRes))
    #weights = []
    #for u in niAna.allUnc:
        #weights.append(1 / u ** 2)
    #waverage = np.average(niAna.allRes, weights=weights)
    #print('Weighted Average:', waverage)
    #sigma = np.std(niAna.allRes)
    #print('Standard deviation:', sigma)
    #plt.errorbar(meas, niAna.allRes, yerr=niAna.allUnc, fmt='bo')
    #plt.plot([0,27],[waverage,waverage], 'r')
    #plt.fill_between([0, 27], waverage - sigma, waverage + sigma, alpha=0.2, linewidth=0,
    #                 color='b')
    #plt.fill_between([0, 27], waverage - sigma - niAna.systUncert, waverage + sigma + niAna.systUncert, alpha=0.2, linewidth=0,
    #                 color='r')
    #plt.show()

    #niAna.calcR(waverage, sigma)

Ana56Ni()
