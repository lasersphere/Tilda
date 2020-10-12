import matplotlib.dates as dates
import math
import os
import sqlite3
import matplotlib.pyplot as plt
import ast
import numpy as np
from Measurement.XMLImporter import XMLImporter
import Physics
import BatchFit
from operator import add
from scipy.optimize import curve_fit
from datetime import datetime
from lxml import etree as et
from XmlOperations import xmlWriteDict
from openpyxl import Workbook, load_workbook


class NiAnalysis:

    def __init__(self, working_dir, db60):
        # working_dir: path of working directory where to save files
        # db: path of database
        # data_path: path where to find xml data files
        # indiv_line_var: line variables to use for fitting
        # indiv_runs60: runs to use for fitting
        # frequ_60ni: literature value of 60Ni transition frequency (Kristian)
        # laser_freq60: laser frequency for 60ni measurements
        # wb: excel workbook containing "good" files
        self.working_dir = working_dir
        self.db = os.path.join(self.working_dir, db60)
        self.data_path = os.path.join(self.working_dir, 'data')
        self.indiv_line_var = ['refLine_0', 'refLine_1', 'refLine_2']
        self.indiv_runs60 = ['AsymVoigt0', 'AsymVoigt1', 'AsymVoigt2']
        self.sum_runs60 = ['AsymVoigtSum0', 'AsymVoigtSum1', 'AsymVoigtSum2']
        self.sum_runs55 = ['AsymVoigt55_0', 'AsymVoigt55_1', 'AsymVoigt55_2', 'AsymVoigt55_All']
        self.tg0 = []
        self.tg1 = []
        self.tg2 = []
        self.frequ_60ni = 850344183
        self.laser_freq60 = 851224124.8007469
        self.laser_freq58 = 851238644.9486578
        self.laser_freq56 = 851253865.030196
        self.laser_freq55 = 851264686.7203143
        self.wb = load_workbook(os.path.join(self.data_path, 'Files.xlsx'))
        try:
            self.results = load_workbook(os.path.join(self.working_dir, 'results\\Isotope_shifts.xlsx'))
        except FileNotFoundError:
            self.results = Workbook()
        self.worksheet = self.results.active
        self.worksheet['A1'] = 'Isotope'
        self.worksheet['B1'] = 'isotope shift'
        self.worksheet['C1'] = 'statistic uncertainty'
        self.worksheet['D1'] = 'method'
        self.worksheet['E1'] = 'date and time'

        # for Calibration
        # times: list of datetimes
        # volts: list of calibrated voltages
        self.times = []
        self.volts = []
        self.fig, self.ax = plt.subplots()

    def start_all(self):
        self.reset()
        self.prep_file_list()
        self.initialize()
        self.asymmetry()
        for run in self.indiv_runs60:
            self.fit_all(niAna.get_files('60Ni'), run)
        self.calib_procedure()

    def ana_individual(self):
        self.isotope_shift('58Ni')
        self.isotope_shift('56Ni')
        self.results.save(os.path.join(self.working_dir, 'results\\Isotope_shifts.xlsx'))

    def reset(self):
        # resets the data base

        # reset fixed shapes
        for run in self.indiv_runs60:
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT fixShape FROM Lines WHERE refRun = ?''', (run,))
            fix_shape = ast.literal_eval(cur.fetchall()[0][0])
            fix_shape['asy'] = [0, 20]
            cur.execute('''UPDATE Lines SET fixShape = ? WHERE refRun = ?''', (str(fix_shape), run,))
            con.commit()
            con.close()

        # reset calibration
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Files SET accVolt = ?''', (29850,))
        con.commit()
        con.close()
        print('Calibration resetted')

        # reset FitResults
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''DELETE FROM FitRes''')
        con.commit()
        con.close()
        print('FitRes cleared')

    def prep_file_list(self):
        # prepares the files database based on the excle workbook

        # create list of "good" files
        ws = self.wb['Tabelle1']
        cells = []
        for i, row in enumerate(ws.rows):
            cells.append([])
            for cell in row:
                cells[i].append(cell.value)
        cells.pop(0)

        # delete "bad" files from database
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT file FROM Files''')
        files = cur.fetchall()
        for f in files:
            remove = 0
            for c in cells:
                if f[0] == c[0] and c[2] == 'y':
                    remove = 0
                    break
                else:
                    remove = 1
            if remove == 1:
                cur.execute('''DELETE FROM Files WHERE file=?''', (f[0],))
        con.commit()
        con.close()
        print('Removed bad files')

    def fit_all(self, files, run):
        # fits all files
        # files: list of files to be fitted
        # run: run to use for fitting

        for f in files:
            # Guess offset
            spec = XMLImporter(path=os.path.join(self.data_path, f))
            offset = (spec.cts[0][0][0] + spec.cts[0][0][-1]) / 2
            self.adj_offset(offset, run)

            # Guess center
            index = list(spec.cts[0][0]).index(max(spec.cts[0][0]))
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (f,))
            acc_v = cur.fetchall()[0][0]
            center_voltage = acc_v - spec.x[0][index]
            cur.execute('''SELECT laserFreq FROM Files WHERE file = ? ''', (f,))
            laser_frequ = cur.fetchall()[0][0]
            cur.execute('''SELECT type FROM Files WHERE file = ?''', (f,))
            iso = cur.fetchall()[0][0]
            if iso == '58Ni':
                mass = 58
            elif iso == '60Ni':
                mass = 60
            else:
                mass = 56
            cur.execute('''SELECT frequency FROM Lines WHERE refRun = ? ''', (run,))
            frequ = cur.fetchall()[0][0]
            v = Physics.relVelocity(Physics.qe * center_voltage, mass * Physics.u)
            v = -v
            center_frequ = Physics.relDoppler(laser_frequ, v) - frequ
            center = center_frequ - 500
            self.adj_center(center, iso)

            BatchFit.batchFit(np.array([f]), self.db, run)

    def adj_offset(self, offset, run):
        # sets offset in database
        # offset: value to set the offset to
        # run: run of which the offset is to be adjusted

        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT shape FROM Lines WHERE refRun LIKE ? ''', (run,))
        shape_dict = ast.literal_eval(cur.fetchall()[0][0])  # create dictionary
        shape_dict['offset'] = offset
        cur.execute('''UPDATE Lines SET shape = ? WHERE refRun = ?''', (str(shape_dict), run,))
        con.commit()
        con.close()
        print('Adjusted Offset to ', shape_dict['offset'])

    def adj_center(self, center, iso):
        # sets center in database
        # center: value to set the center to
        # iso: istope of which the center is to be adjsued

        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso = ?''', (center, iso,))
        con.commit()
        con.close()
        print('Adjusted Center to', center)

    def initialize(self):
        # first fit of reference measurements (60Ni)

        files = self.get_files('60Ni')

        # find software gates:
        self.set_timegates(files)

        for run in self.indiv_runs60:
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            if '0' in run:
                t_mid = self.tg0[0] + (self.tg0[1] - self.tg0[0]) / 2
                cur.execute('''UPDATE Isotopes SET midTof = ? WHERE iso = ?''', (t_mid, '60Ni',))
                cur.execute('''UPDATE Runs SET softwGateWidth = ? WHERE run = ?''',
                            (self.tg0[1] - self.tg0[0], 'AsymVoigt0'))
            elif '1' in run:
                t_mid = self.tg1[0] + (self.tg1[1] - self.tg1[0]) / 2
                cur.execute('''UPDATE Isotopes SET midTof = ? WHERE iso = ?''', (t_mid, '60Ni',))
                cur.execute('''UPDATE Runs SET softwGateWidth = ? WHERE run = ?''',
                            (self.tg1[1] - self.tg1[0], 'AsymVoigt1'))
            else:
                t_mid = self.tg2[0] + (self.tg2[1] - self.tg2[0]) / 2
                cur.execute('''UPDATE Isotopes SET midTof = ? WHERE iso = ?''', (t_mid, '60Ni',))
                cur.execute('''UPDATE Runs SET softwGateWidth = ? WHERE run = ?''',
                            (self.tg2[1] - self.tg2[0], 'AsymVoigt2'))
            con.commit()
            con.close()

            self.fit_all(files, run)
        plt.close()  # without this, some pyplot error is thrown...

    def asymmetry(self):
        # find fix asymmetry parameter

        con = sqlite3.connect(self.db)
        cur = con.cursor()
        files = self.get_files('60Ni')
        for run in self.indiv_runs60:
            asy = []
            unc = []
            weights = []

            # collect asymmetry parameter for each file
            for f in files:
                print(f)
                cur.execute('''SELECT pars FROM FitRes WHERE file = ? AND run = ?''', (f, run,))
                pars = ast.literal_eval(cur.fetchall()[0][0])['asy']    # create dictionary
                asy.append(pars[0])
                unc.append(pars[1])
                weights.append(1 / (pars[1] ** 2))

            # calculate weighted mean of all parameters
            mean = np.average(asy, weights=weights)
            std = np.std(asy)

            # plot results
            fig, ax = plt.subplots()
            ax.errorbar(list(range(0, len(asy))), asy, yerr=unc)
            ax.plot([0, len(asy)], [mean, mean], 'r-')
            ax.fill_between([0, len(asy)], [mean + std, mean + std], [mean - std, mean - std], color='r', alpha=0.2)
            ax.set_title('Run: ' + run)
            ax.set_xlabel('Messungen')
            ax.set_ylabel('Asymmetry factor')
            ax.text(20, 13, str(mean))
            #fig.show()  # show plot
            fig.savefig(os.path.join(self.working_dir, 'asymmetry\\Scaler_' + str(run) + '.png'))  # save plot
            plt.close()

            # set asymmetry parameter to mean and fix it
            cur.execute('''SELECT shape FROM Lines WHERE refRun = ?''', (run,))
            shape = ast.literal_eval(cur.fetchall()[0][0])  # create dictionary
            cur.execute('''SELECT fixShape FROM Lines WHERE refRun = ?''', (run,))
            fix = ast.literal_eval(cur.fetchall()[0][0])    # create dictionary
            shape['asy'] = mean
            fix['asy'] = True
            cur.execute('''UPDATE Lines SET shape = ? WHERE refRun = ?''', (str(shape), run,))
            cur.execute('''UPDATE Lines SET fixShape = ? WHERE refRun = ?''', (str(fix), run,))
            con.commit()
        con.close()

    def get_files(self, isotope):
        # return list of all files in database of one isotope
        # isotope: istope of which to get the files of

        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT file FROM Files WHERE type LIKE ? ''', (isotope,))
        files = cur.fetchall()
        con.close()
        return [f[0] for f in files]

    def calibrate_all_ref(self):
        # Calibrate DAC voltage of all refernce files on transition frequency of 60Ni (Kristian)

        files = self.get_files('60Ni')
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        tuples = []
        for f in files:
            cur.execute('''SELECT pars FROM FitRes WHERE file = ?''', (f,))
            pars = cur. fetchall()
            center = []
            errs = []
            weights = []

            # each file has fit results for 3 different runs (3 scalers)
            for result in pars:
                dic = ast.literal_eval(result[0])
                center.append(dic['center'][0])
                errs.append(dic['center'][1])
                weights.append(1 / (dic['center'][1] ** 2))

            # use mean of 3 scalers for calibration of each file
            mean = np.average(center, weights=weights)
            acc_volt = self.calibrate(f, mean)
            cur.execute('''UPDATE Files Set accVolt = ? WHERE file = ?''', (acc_volt, f,))
        con.commit()
        con.close()
        self.plot_calibration()

    def calibrate(self, file, delta_frequ):
        # returns calibrated voltage of a single reference file based on the mean fitresult of the center - parameter
        # file: file to calibrate
        # delta_frequ: difference of fitted frequence and literature value

        # Calculate differential Doppler shift
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT accVolt from Files WHERE file = ?''', (file,))
        acc_volt = cur.fetchall()[0][0]
        con.close()
        diff_dopp60 = Physics.diffDoppler(delta_frequ + self.frequ_60ni, acc_volt, 60)

        # Calculate calibration
        delta_u = delta_frequ / diff_dopp60
        return acc_volt + delta_u

    def plot_calibration(self):
        # plots calibrated voltage of all reference files in the database over the time

        files = self.get_files('60Ni')

        files.sort()

        file_times = []
        voltages = []
        for f in files:
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT date FROM Files  WHERE file = ?''', (f,))
            date = cur.fetchall()[0][0]
            cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (f,))
            acc_vol = cur.fetchall()[0][0]
            con.close()
            file_times.append(datetime.strptime(date, '%Y-%m-%d %H:%M:%S'))  # convert string to datetime
            voltages.append(acc_vol)
        self.times = file_times
        self.volts = voltages

        plt.close()
        self.fig, self.ax = plt.subplots()
        self.ax.plot_date(file_times, voltages, fmt='-b')
        self.ax.xaxis.set_major_formatter(dates.DateFormatter('%Y-%m-%d %H:%M:%S'))
        plt.setp(self.ax.get_xticklabels(), rotation=30, ha="right")
        self.ax.set_title('Calibrated Voltage')
        self.ax.set_ylabel('Voltage in V')

    def calibrate_all(self, files, color):
        # use calibration of reference measurements for calibration of all files
        # files: files to use calibration on

        times = []
        volts = []
        con = sqlite3.connect(self.db)
        cur = con.cursor()

        # select time stamp of file and use interpolation to find calibrated voltage
        for f in files:
            cur.execute('''SELECT date FROM Files WHERE file = ?''', (f,))
            time = datetime.strptime(cur.fetchall()[0][0], '%Y-%m-%d %H:%M:%S')
            times.append(time)
            if time <= self.times[0]:
                volts.append(self.volts[0])
                cur.execute('''UPDATE Files SET accVolt = ? WHERE file = ?''', (self.volts[0], f,))
                continue
            for i, t in enumerate(self.times):
                if i < len(self.times) - 1:
                    if t <= time < self.times[i + 1]:
                        y0, m = curve_fit(self.lin_func, dates.date2num(self.times[i:i+2]), self.volts[i:i+2])[0]
                        new_volt = y0 + m * dates.date2num(time)
                        volts.append(new_volt)
                        cur.execute('''UPDATE Files SET accVolt = ? WHERE file = ?''', (new_volt, f,))
                        break
                else:
                    if t <= time:
                        volts.append(self.volts[-1])
                        cur.execute('''UPDATE Files SET accVolt = ? WHERE file = ?''', (self.volts[-1], f,))
        con.commit()
        con.close()
        self.ax.plot_date(times, volts, fmt=color)
        #self.fig.show()
        self.fig.savefig(os.path.join(self.working_dir, 'calibration\\Voltagecalibration.png'))

    def calib_procedure(self):
        # the whole procedure of calibration
        # isotpe: isotpe to use the calibration on

        plt.close()  # without this, some pyplot error is thrown...

        # calibration of reference files
        self.calibrate_all_ref()

        # refit with new voltage
        for run in self.indiv_runs60:
            self.fit_all(self.get_files('60Ni'), run)
        plt.close()  # without this, some pyplot error is thrown...

        # calibration of reference files (second time for higher accurancy)
        self.calibrate_all_ref()

        # calibrate all other files
        self.calibrate_all(niAna.get_files('58Ni'), 'rx')
        self.calibrate_all(niAna.get_files('56Ni'), 'yx')
        self.calibrate_all(niAna.get_files('55Ni'), 'gx')

        # refit reference files
        for run in self.indiv_runs60:
            self.fit_all(self.get_files('60Ni'), run)

    @staticmethod
    def lin_func(x, y0, m):
        # linear function

        return y0 + m * x

    def center(self, isotope):
        # analysis procedure plots results of each file and returns the weighted mean and standard deviation of all

        files = self.get_files(isotope)

        # if isotope is not reference, fit first. Reference is already fitted
        if isotope != '60Ni':

            # find software gates:
            self.set_timegates(files)

            for run in self.indiv_runs60:
                con = sqlite3.connect(self.db)
                cur = con.cursor()
                if '0' in run:
                    t_mid = self.tg0[0] + (self.tg0[1] - self.tg0[0]) / 2
                    cur.execute('''UPDATE Isotopes SET midTof = ? WHERE iso = ?''', (t_mid, isotope,))
                    cur.execute('''UPDATE Runs SET softwGateWidth = ? WHERE run = ?''',
                                (self.tg0[1] - self.tg0[0], 'AsymVoigt0'))
                elif '1' in run:
                    t_mid = self.tg1[0] + (self.tg1[1] - self.tg1[0]) / 2
                    cur.execute('''UPDATE Isotopes SET midTof = ? WHERE iso = ?''', (t_mid, isotope,))
                    cur.execute('''UPDATE Runs SET softwGateWidth = ? WHERE run = ?''',
                                (self.tg1[1] - self.tg1[0], 'AsymVoigt1'))
                else:
                    t_mid = self.tg2[0] + (self.tg2[1] - self.tg2[0]) / 2
                    cur.execute('''UPDATE Isotopes SET midTof = ? WHERE iso = ?''', (t_mid, isotope,))
                    cur.execute('''UPDATE Runs SET softwGateWidth = ? WHERE run = ?''',
                                (self.tg2[1] - self.tg2[0], 'AsymVoigt2'))
                con.commit()
                con.close()
                self.fit_all(files, run)
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        times = []
        centers = []
        errs = []
        for f in files:
            cur.execute('''SELECT pars FROM FitRes WHERE file = ?''', (f,))
            pars = cur.fetchall()
            results = []
            weihgts = []
            # there are three results for each file (three scalers)
            for run in pars:
                results.append(ast.literal_eval(run[0])['center'][0])
                weihgts.append(1 / (ast.literal_eval(run[0])['center'][1] ** 2))
            # take mean of the three scalers
            mean = np.average(results, weights=weihgts)
            std = np.std(results)
            centers.append(mean)
            errs.append(std)
            cur.execute('''SELECT date FROM Files WHERE file = ?''', (f,))
            date = datetime.strptime(cur.fetchall()[0][0], '%Y-%m-%d %H:%M:%S')
            times.append(date)
        con.close()
        plt.close()  # without this, some pyplot error is thrown...
        upper = []  # for plotting
        lower = []  # for plotting
        for i, c in enumerate(centers):
            upper.append(c + errs[i])
            lower.append(c - errs[i])
        fig, ax = plt.subplots()
        ax.plot_date(times, centers, fmt='bx', linestyle='-', )
        plt.setp(ax.get_xticklabels(), rotation=30, ha="right")
        ax.fill_between(times, upper, lower, alpha=0.2)
        ax.set_title('Center frequency of ' + isotope)
        ax.set_ylabel('Frequency in MHz')
        #fig.show()
        fig.savefig(os.path.join(self.working_dir, 'center\\'+ isotope + '.png'))
        return np.average(centers, weights=errs), np.std(centers)  # mean of all files is returned

    def isotope_shift(self, isotope):
        # calculate isotope shift of isotope using fitted centers
        # isotope: isotope to calculate the isotope shift of

        ref_center, ref_unc = self.center('60Ni')
        iso_center, iso_unc = self.center(isotope)
        isotope_shift = iso_center - ref_center
        ishift_unc = np.sqrt(np.square(iso_unc) + np.square(ref_unc))
        self.worksheet.append([isotope, isotope_shift, ishift_unc, 'individual',
                               datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        print(isotope_shift, ishift_unc)

    def t_proj(self, file, scaler):
        # returns the timeprojection of file as a list
        # file: file to return the timeprojection of
        # scaler: scaler to return the timeprojection of
        spec = XMLImporter(path=os.path.join(self.working_dir, 'data\\' + file))
        return spec.t_proj[0][scaler]

    @staticmethod
    def gauss(time, cts, sigma, t_mid, off):
        # returns a gaussian

        return off + cts / np.sqrt(2 * np.pi * sigma ** 2) * np.exp(-1 / 2 * np.square((time - t_mid) / sigma))

    @staticmethod
    def lorentz(time, cts, gamma, t_mid, off):
        # returns a lorentzian

        lw = gamma * 2
        return off + (cts * 2 / (math.pi * lw)) * ((lw ** 2 / 4) / ((time - t_mid) ** 2 + (lw ** 2 / 4)))

    def fit_time(self, time, cts):
        # fits a gaussian or lorentzian to a time projection
        # time: time axis
        # cts: list of counts

        # guess Start parameters and set amplitude and sigma positive
        # amplitude as max(cts), gamma as 10, center as position of max(cts), offset as mean of first and last cts
        start_par = np.array([max(cts), 10, time[cts.index(max(cts))], (time[0] + time[-1]) / 2])
        param_bounds = ([0, 0, -np.inf, -np.inf], [np.inf, np.inf, np.inf, np.inf])

        # Fitting
        # gauss
        # a, sigma, center, offset = curve_fit(self.gauss, time, cts, start_par, bounds=param_bounds)[0]
        # return a, sigma, center, offset
        #lorentzian
        a, gamma, center, offset = curve_fit(self.lorentz, time, cts, start_par, bounds=param_bounds)[0]
        return a, gamma, center, offset

    def find_timegates(self, files, scaler):
        # returns center and width of timegate
        # files: list of files to use for timegate determination
        # scaler: scaler to use for timegate determination

        # create a list to store counts for each timestep in
        t_proj = np.zeros(1024)

        # sum up counts for each timestep of all files

        for f in files:
            # for each element of t_proj: add counts of file to t_proj
            t_proj = list(map(add, t_proj, self.t_proj(f, scaler)))

        # create list of timesteps for fitting and plotting
        time = list(range(0, len(t_proj)))

        # fit a gaussian or lorentzian
        a, sigma, center, offset = self.fit_time(time, t_proj)

        # Plot counts and fit and a 1-sigma timegate
        fig, ax = plt.subplots()
        ax.plot(time, t_proj, 'b.')
        ax.set_title('Scaler' + str(scaler))
        ax.plot(time, self.lorentz(time, a, sigma, center, offset), 'r-')
        ax.axvline(center - 1 * sigma, color='y')
        ax.axvline(center + 1 * sigma, color='y')
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT type FROM Files WHERE file = ?''', (files[0],))
        isotope = cur.fetchall()[0][0]
        con.close()
        fig.savefig(os.path.join(self.working_dir, 't_proj\\' + isotope + str(scaler) + '.png'))
        return center, sigma

    def set_timegates(self, files):

        print('Setting timegates to:')
        # find software gates:
        for s in [0, 1, 2]:
            t_mid, t_width = self.find_timegates(files, s)
            multiple = 1
            t_min = (t_mid - multiple * t_width) / 100  # divide by 100 to convert bins to microseconds
            t_max = (t_mid + multiple * t_width) / 100  # divide by 100 to convert bins to microseconds
            print('Scaler', s, ':', t_min, t_max)
            if s == 0:
                self.tg0 = [t_min, t_max]
            if s == 1:
                self.tg1 = [t_min, t_max]
            if s == 2:
                self.tg2 = [t_min, t_max]

    def stack_files(self, isotope):
        print('Stacking files of isotope', isotope)
        files = self.get_files(isotope)
        scalers = [0, 1, 2]

        sumcts = [[], [], []]
        voltage = [[], [], []]
        binsizes = []

        # iterate through scalers:
        for s in scalers:
            # find and set timegates
            t_mid, t_width = self.find_timegates(files, s)
            multiple = 2
            t_min = (t_mid - multiple * t_width) / 100  # divide by 100 to convert bins to microseconds
            t_max = (t_mid + multiple * t_width) / 100  # divide by 100 to convert bins to microseconds

            fig, ax = plt.subplots()
            volt_cts = [] # Will be a list of tuples: (DAC - offset, cts, bg)

            # iterate through files to sum up counts in timegate
            for f in files:
                # spectrum only in the specified time gate
                spec = XMLImporter(path=self.working_dir + '\\data\\' + str(f),
                                   softw_gates=[[-35, 15, t_min, t_max], [-35, 15, t_min, t_max],
                                                [-35, 15, t_min, t_max]])
                # spectrum of background
                bg = XMLImporter(path=self.working_dir + '\\data\\' + str(f),
                                 softw_gates=[[-35, 15, 0.5, 4], [-35, 15, 0.5, 4],
                                              [-35, 15, 0.5, 4]])

                #normalization of background to number of bins
                norm_factor = 2 * multiple * t_width / 100 / 3.5  # 2 * multiple is the total timegate width
                for i, c in enumerate(bg.cts[0][s]):
                    bg.cts[0][s][i] = c * norm_factor

                # use calibration
                for j, x in enumerate(spec.x[0]):
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''SELECT accVolt from Files WHERE file  = ?''', (f,))
                    accV = cur.fetchall()[0][0]
                    con.close()
                    offset = (accV - 29850)
                    volt_cts.append((x - offset, spec.cts[0][s][j], bg.cts[0][s][j]))

                # plot uncalibrated spectrum
                plt.close('all')
                ax.plot(spec.x[0], spec.cts[0][s])
                ax.plot(bg.x[0], bg.cts[0][s])
            ax.set_title(isotope + ' Scaler ' + str(s) + ': raw data')
            ax.set_xlabel('DAC Voltage in V')
            ax.set_ylabel('Counts')
            fig.savefig(os.path.join(self.working_dir, 'raw_spectra\\' + isotope + str(s) + '.png'))
            #fig.show()

            volt_cts.sort()

            # find all entries belonging to one voltage bin and sum up the counts and background
            voltages = []
            weights = []
            c_sum = 0
            b_sum = 0
            binned_volt_cts = []
            v_min = volt_cts[0][0]
            for tup in volt_cts:
                if tup[0] < v_min + 0.9:
                    voltages.append(tup[0])
                    weights.append(tup[1])
                    c_sum += tup[1]
                    b_sum += tup[2]
                else:
                    v_min = tup[0]
                    if weights == [0]:
                        binned_volt_cts.append((voltages[0], c_sum, b_sum))
                    elif weights == []:
                        continue
                    else:
                        binned_volt_cts.append((np.average(voltages, weights=weights), c_sum, b_sum))
                    voltages = []
                    weights = []
                    c_sum = 0
                    b_sum = 0

            # Plot voltage bins:
            # create voltage axis:
            voltages = []
            for tup in binned_volt_cts:
                voltages.append(tup[0])
            # create step axis
            steps = list(range(0, len(voltages)))
            # plot
            fig, ax = plt.subplots()
            ax.plot(steps, voltages, 'b.')
            ax.set_xlabel('step')
            ax.set_ylabel('Voltage in V')
            ax.set_title('DAC voltage bins')

            # find linear fit. slope is the binsize
            model = np.polyfit(steps, voltages, 1)
            predict = np.poly1d(model)
            binsizes.append(model[0])
            voltages = predict(steps)
            ax.plot(steps, voltages, 'r-')
            fig.savefig(os.path.join(self.working_dir, 'voltage_steps\\' + isotope + str(s) + '.png'))
            plt.close()

            # assign new voltage bins
            binned_volt_cts_0 = []
            for i, tup in enumerate(binned_volt_cts):
                #binned_volt_cts[i] = (predict(i), tup[1], tup[2])
                binned_volt_cts_0.append((predict(i), tup[1], tup[2]))

            i = 0
            j = 0
            voltages = []
            binned_volt_cts = []
            for tup in binned_volt_cts_0:
                i += 1
                if i == 1:
                    v = tup[0]
                    binned_volt_cts.append(tup)
                    voltages.append(v)
                else:
                    binned_volt_cts[j] = (v, binned_volt_cts[j][1] + tup[1], binned_volt_cts[j][2] + tup[2])
                if i == 2:
                    j += 1
                    i = 0

            # plot calibrated and summed data
            # create counts axis
            c_sum = [tup[1] for tup in binned_volt_cts]
            b_sum = [tup[2] for tup in binned_volt_cts]
            # plot
            fig, ax = plt.subplots()
            ax.plot(voltages, c_sum, 'b.')
            ax.plot(voltages, b_sum, ' r.')
            ax.set_title('Calibrated and summed Spectrum of ' + isotope + ', Scaler ' + str(s))
            ax.set_xlabel('DAC Voltage in V')
            ax.set_ylabel('Counts')
            fig.savefig(os.path.join(self.working_dir, 'spectra\\calibrated\\' + isotope + str(s) + '.png'))
            #fig.show()

            # calculate statistic uncertainty
            unc = []
            for cts in c_sum:
                unc.append(np.sqrt(cts))

            # normalize
            c_sum_norm = []
            unc_norm = []
            for i, cts in enumerate(c_sum):
                if b_sum[i] == 0:
                    c_sum_norm.append(0)
                    unc_norm.append(0)
                else:
                    c_sum_norm.append(int((cts) / b_sum[i] * np.mean(b_sum)))
                    unc_norm.append(int((unc[i] / b_sum[i] * np.mean(b_sum))))
            # plot normalized spectrum
            fig, ax = plt.subplots()
            ax.errorbar(voltages, c_sum_norm, yerr=unc_norm, fmt='b.')
            ax.set_title('Calibrated, summed and normalized. Scaler' + str(s))
            ax.set_xlabel('DAC Voltage in V')
            ax.set_ylabel('Counts')
            fig.savefig(os.path.join(self.working_dir, 'spectra\\normalized\\' + isotope + str(s) + '.png'))

            # assign
            c_sum_norm = c_sum_norm[:-5]
            voltages = voltages[:-5]
            sumcts[s] = c_sum_norm
            voltage[s] = voltages

        # creating file:

        # calculate mean binsize:
        binsize = np.average(binsizes) * 2

        # prepare scaler array for xml-file
        scaler_array = []
        for s in scalers:
            timestep = 0
            for i, c in enumerate(sumcts[s]):
                scaler_array.append((s, i, timestep, c))
                timestep += 1

        # Create dictionary for xml export
        if isotope == '60Ni':
            laser_freq = self.laser_freq60
        elif isotope == '58Ni':
            laser_freq = self.laser_freq58
        elif isotope == '56Ni':
            laser_freq = self.laser_freq56
        elif isotope == '55Ni':
            laser_freq = self.laser_freq55
        file_creation_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        header_dict = {'type': 'trs',
                       'isotope': isotope + '_sum',
                       'isotopeStartTime': file_creation_time,
                       'accVolt': 29850,
                       'laserFreq': Physics.wavenumber(laser_freq) / 2,
                       'nOfTracks': 1,
                       'version': 99.0}
        track0_dict_header = {'trigger': {},  # Need a trigger dict!
                              'activePmtList': [0, 1, 2],  # Must be in form [0,1,2]
                              'colDirTrue': True,
                              'dacStartRegister18Bit': 0,
                              'dacStartVoltage': voltage[0][0],
                              'dacStepSize18Bit': None,  # old format xml importer checks whether val or None
                              'dacStepsizeVoltage': binsize,
                              'dacStopRegister18Bit': len(voltage[0]) - 1,  # not real but should do the trick
                              'dacStopVoltage': voltage[0][-1],
                              'invertScan': False,
                              'nOfBins': len(voltage[0]),
                              'nOfCompletedSteps': float(len(voltage[0])),
                              'nOfScans': 1,
                              'nOfSteps': len(voltage[0]),
                              'postAccOffsetVolt': 0,
                              'postAccOffsetVoltControl': 0,
                              'softwGates': [[-40, -10, 0, timestep], [-40, -10, 0, timestep], [-40, -10, 0, timestep]],
                              # 'softwGates': [[-252, -42, 0, 0.4], [-252, -42, 0, 0.4], [-252, -42, 0, 0.4]],
                              # For each Scaler: [DAC_Start_Volt, DAC_Stop_Volt, scaler_delay, softw_Gate_width]
                              'workingTime': [file_creation_time, file_creation_time],
                              'waitAfterReset1us': 0,  # looks like I need those for the importer
                              'waitForKepco1us': 0  # looks like I need this too
                              }
        data = '['
        for i, s in enumerate(scaler_array):
            data = data + str(scaler_array[i]) + ' '
        data = data[:len(data) - 1]
        data = data + ']'
        track0_dict_data = {
            'scalerArray_explanation': 'continously acquired data. List of Lists, each list represents the counts of '
                                       'one scaler as listed in activePmtList.Dimensions are: (len(activePmtList), '
                                       'nOfSteps), datatype: np.int32',
            'scalerArray': data}
        track0_vol_proj = {'voltage_projection': np.array(sumcts),
                           'voltage_projection_explanation': 'voltage_projection of the time resolved data. List of '
                                                             'Lists, each list represents the counts of one scaler as '
                                                             'listed in activePmtList.Dimensions are: '
                                                             '(len(activePmtList), nOfSteps), datatype: np.int32'}
        dictionary = {'header': header_dict,
                      'tracks': {'track0': {'header': track0_dict_header,
                                            'data': track0_dict_data,
                                            'projections': track0_vol_proj
                                            },
                                 }
                      }

        # Write xml-file
        root = et.Element('BecolaData')

        xmlWriteDict(root, dictionary)
        xml = et.ElementTree(root)
        xml.write(self.working_dir + '\\data\\BECOLA_Stacked' + isotope + '.xml')

        # Add to database
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''INSERT OR IGNORE INTO Files (file, filePath, date, type) VALUES (?, ?, ?, ?)''',
                    ('BECOLA_Stacked' + isotope + '.xml', 'data\BECOLA_Stacked' + isotope + '.xml',
                     file_creation_time, isotope + '_sum'))
        cur.execute('''UPDATE Files SET offset = ?, accVolt = ?,  laserFreq = ?, laserFreq_d = ?, colDirTrue = ?, 
            voltDivRatio = ?, lineMult = ?, lineOffset = ?, errDateInS = ? WHERE file = ? ''',
            ('[0]', 29850, laser_freq, 0, True, str({'accVolt': 1.0, 'offset': 1.0}), 1, 0,
             1, 'BECOLA_Stacked' + isotope + '.xml'))
        con.commit()
        con.close()

    def fit_stacked(self, isotope, run, sym=True):
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        print(run)
        cur.execute('''SELECT fixShape from Lines WHERE refRun = ? ''', (run,))
        shape = cur.fetchall()
        shape_dict = ast.literal_eval(shape[0][0])
        con.close()
        if sym:
            shape_dict['asy'] = True
        else:
            shape_dict['asy'] = [0,30]
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Lines SET fixShape = ? WHERE refRun = ?''', (str(shape_dict), run))
        con.commit()
        con.close()


        BatchFit.batchFit(['BECOLA_Stacked' + isotope + '.xml'], self.db, run, x_as_voltage=True,
                          save_file_as='.png')

    def center_stacked(self, isotope):
        self.stack_files(isotope)
        runs = self.sum_runs60
        if isotope == '55Ni':
            runs = self.sum_runs55
        for run in runs:
            self.fit_stacked(isotope, run)
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT pars FROM FitRes WHERE iso = ?''', (isotope + '_sum',))
        pars = cur.fetchall()
        results = []
        weihgts = []
        # there are three results for each file (three scalers)
        for run in pars:
            results.append(ast.literal_eval(run[0])['center'][0])
            weihgts.append(1 / (ast.literal_eval(run[0])['center'][1] ** 2))
        # take mean of the three scalers
        mean = np.average(results, weights=weihgts)
        std = np.std(results)
        return mean, std

    def isotope_shift_stacked(self, isotope):
        # calculate isotope shift of isotope using fitted centers of stacked file
        # isotope: isotope to calculate the isotope shift of

        ref_center, ref_unc = self.center_stacked('60Ni')
        plt.close('all')
        iso_center, iso_unc = self.center_stacked(isotope)

        isotope_shift = iso_center - ref_center
        ishift_unc = np.sqrt(np.square(iso_unc) + np.square(ref_unc))
        self.worksheet.append([isotope, isotope_shift, ishift_unc, 'stacked',
                               datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        print(isotope_shift, ishift_unc)



# workingDir = 'D:\\Owncloud\\User\\Laura\\Nickelauswertung'  # Path Laptop
workingDir = 'C:\\Users\\Laura Renth\\ownCloud\\User\\Laura\\Nickelauswertung'  # Path IKP
db = 'Nickel_BECOLA_60Ni.sqlite'

niAna = NiAnalysis(workingDir, db)
#niAna.start_all()
#niAna.ana_individual()

niAna.isotope_shift_stacked('55Ni')
niAna.results.save(os.path.join(niAna.working_dir, 'results\\Isotope_shifts.xlsx'))