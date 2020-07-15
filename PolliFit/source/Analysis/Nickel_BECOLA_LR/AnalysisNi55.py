from Measurement.XMLImporter import XMLImporter
from openpyxl import Workbook
# import Tools
import os
import ast
import matplotlib.pyplot as plt
import numpy as np
import sqlite3
import BatchFit
import Physics
from operator import add
from scipy.optimize import curve_fit
from datetime import datetime
from XmlOperations import xmlWriteDict
from lxml import etree as ET



# from scipy import constants as const


class NiAnalysis:
    def __init__(self, cal_t, cal_g, syst_u):
        # parameters:   cal_t: list of tuples. Each tuple is a pair of one 58Ni and one 60Ni
        #               cal_g: list of tuples. Each tuple is a pair of one 58Ni and one 56Ni
        #               syst_u:
        self.lastRow = 1
        self.workingDir = 'D:\\Daten\\IKP\\Nickel-Auswertung\\Auswertung'
        self.filepath = os.path.join(self.workingDir, 'data')
        self.db = os.path.join(self.workingDir, 'Nickel_BECOLA.sqlite')
        self.db_stacked = os.path.join(self.workingDir, 'Nickel_BECOLAstacked.sqlite')
        print('Database:', self.db)
        self.runs = ['AsymVoigt0', 'AsymVoigt1', 'AsymVoigt2']
        self.run = ''
        self.linVars = ['58_0', '58_1', '58_2']
        self.lineVar = ''
        self.files = ['Scaler0.xlsx', 'Scaler1.xlsx', 'Scaler2.xlsx']
        self.xlsFile = ''
        self.laserFreq55 = 851264686.7203143
        self.xlsFileAll = self.workingDir + '\\AllPmts.xlsx'
        self.wbAll = Workbook()
        # Create a worksheet for calibration data
        self.wsCalAll = self.wbAll.create_sheet('Calibration', 0)
        self.wsCalAll.title = 'Calibration'
        self.wsCalAll.cell(row=self.lastRow, column=1, value='Tuple')
        self.wsCalAll.cell(row=self.lastRow, column=2, value='old Voltage')
        self.wsCalAll.cell(row=self.lastRow, column=3, value='old isotope shift')
        self.wsCalAll.cell(row=self.lastRow, column=4, value='calibrated Voltage')
        self.wsCalAll.cell(row=self.lastRow, column=5, value='calibrated istope shift')
        # Create a worksheet for isotope shift data
        self.wsISAll = self.wbAll.create_sheet('Isotope shifts 56Ni', 1)
        self.wsISAll.title = 'Isotope shifts 56Ni'
        self.wsISAll.cell(row=1, column=1, value='File')
        self.wsISAll.cell(row=1, column=2, value='Isotope shift')
        self.wsISAll.cell(row=1, column=3, value='Uncertainty')
        self.wbAll.save(self.xlsFileAll)

        self.calTuples = cal_t
        self.calGroups56 = cal_g
        self.shifts56 = []
        self.uncert56 = []
        self.systUncert = syst_u
        self.allRes = []
        self.allUnc = []
        self.allMeas = []

    def prep(self):
        # Calibrate for all pmts together
        self.calibrate()
        print('-------------------- Calibration done')
        self.assign_cal()
        print('-------------------- Assignement done')
        # evaluate for each scaler individually
        #for i, run in enumerate(self.runs):
            #self.run = run
            #self.lineVar = self.linVars[i]

            # Fit Ni 56
            #file_list56 = self.get_files('56Ni')
            #self.fit_all(file_list56)

    def reset(self):
        # reset calibration
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Files SET accVolt = ?''', (29850,))
        con.commit()
        con.close()

    def calibrate(self):

        # Adjust center in isotopes
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (-200, '58Ni',))
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (300, '60Ni',))
        con.commit()
        con.close()

        # Find files to fit
        file_list58 = self.get_files('58Ni')
        for i,run in enumerate(self.runs):
            self.run = run
            self.lineVar = self.linVars[i]
            self.fit_all(file_list58)
        file_list60 = self.get_files('60Ni')
        for i, run in enumerate(self.runs):
            self.run = run
            self.lineVar = self.linVars[i]
            self.fit_all(file_list60)

        # Calibrate
        shifts, files, uncert = self.single_cali()

        # Plot the uncalibrated isotope shift of 60Ni
        plt.errorbar(files, shifts, yerr=uncert, fmt='bo')
        plt.title('Uncalibrated Reference Shift')
        plt.show()

        # Adjust center in isotopes
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (-400, '58Ni',))
        cur.execute('''UPDATE Isotopes SET center = ? WHERE iso LIKE ?''', (100, '60Ni',))
        con.commit()
        con.close()

        # Fit again after calibration
        for i, run in enumerate(self.runs):
            self.run = run
            self.lineVar = self.linVars[i]
            self.fit_all(file_list58)
        for i, run in enumerate(self.runs):
            self.run = run
            self.lineVar = self.linVars[i]
            self.fit_all(file_list60)
        shifts = []
        uncert = []
        for tup in self.calTuples:
            t_ishifts = []
            t_unc = []
            for i, run in enumerate(self.runs):
                self.run = run
                self.lineVar = self.linVars[i]
                ishift, err = self.cal_ishift(tup)
                t_ishifts.append(ishift)
                t_unc.append(err)
            weights = []
            for u in t_unc:
                weights.append(1 / u ** 2)
            waverage = np.average(t_ishifts, weights=weights)
            sigma = np.std(t_ishifts)
            print('Calibrated isotope shift of File', tup[0], 'is', waverage)
            shifts.append(waverage)
            uncert.append(sigma)
        plt.errorbar(files, shifts, yerr=uncert, fmt='bo')
        plt.title('Calibrated Reference Shift')
        plt.show()

        # Save calibrated istotope shifts to excel
        r = 2
        for s in shifts:
            self.wsCalAll.cell(row=r, column=5, value=s)
            r += 1
        self.wbAll.save(self.xlsFileAll)

    def assign_cal(self):
        # Assign calibration to 56Ni files
        for tup in self.calGroups56:
            print('Tuple to assigne to :', tup[1])
            for file in tup[1]:
                print('File:', file)
                file56 = 'BECOLA_' + str(file) + '.xml'
                file580 = 'BECOLA_' + str(tup[0][0]) + '.xml'
                file581 = 'BECOLA_' + str(tup[0][1]) + '.xml'
                print(file581)

                # Query calibrated voltage from 58Ni
                con = sqlite3.connect(self.db)
                cur = con.cursor()
                cur.execute('''SELECT accVolt FROM files WHERE file LIKE ?''', (file580,))
                cal_volt = cur.fetchall()[0][0]
                cur.execute('''SELECT accVolt FROM files WHERE file LIKE ?''', (file581,))
                cal_volt = (cal_volt + cur.fetchall()[0][0]) / 2
                con.close()

                # Update 56Ni voltage to calibration
                con = sqlite3.connect(self.db)
                cur = con.cursor()
                cur.execute('''UPDATE Files SET accVolt = ? WHERE file LIKE ?''', (cal_volt, file56,))
                con.commit()
                con.close()

                print('Voltage updated for file', file56)

    def get_files(self, iso):
        # param:    iso: isotope to fetch files of
        # returns a list of files
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT file FROM Files WHERE type LIKE ? ''', (iso,))
        files = cur.fetchall()
        con.close()
        return [f[0] for f in files]

    def fit_all(self, file_list):
        for file in file_list:
            print('Starting File', file)
            # asymmetry factor for 58Ni and isotope of interest should be the same:
            for tup in self.calTuples:
                if 'BECOLA_' + str(tup[1]) + '.xml' == file:
                    file58 = 'BECOLA_' + str(tup[0]) + '.xml'
                    # Fetch asymmetry factor from 58Ni
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''SELECT pars from FitRes WHERE file = ? AND run = ?''', (file58, self.run))
                    pars = cur.fetchall()
                    pars = ast.literal_eval(pars[0][0])
                    asy = pars['asy'][0]
                    print('The Asymmetry Factor of file', file58, ' and run', self.run, 'is', asy)
                    con.close()
                    # Update asymmetry factor of 60Ni
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET shape = ?''', (
                        "{'name': 'VoigtAsy', 'sigma': 200, 'offset': 343.0, 'asy': " + str(asy) +
                        " , 'gamma': 20,'offsetSlope': 0}",))
                    con.commit()
                    con.close()
                    # fix asymmetry factor
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET fixShape = ?''', (
                        "{'asy': True, 'offsetSlope': True, 'sigma': [0,200], 'offset': False, 'gamma': [0,50]}",))
                    con.commit()
                    con.close()
            for tup in self.calGroups56:
                if 'BECOLA_' + str(tup[1]) + '.xml' == file:
                    file58 = 'BECOLA_' + str(tup[0]) + '.xml'
                    # Fetch asymmetry factor of 58Ni
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''SELECT pars from FitRes WHERE file = ? AND run = ?''', (file58, self.run))
                    pars = cur.fetchall()
                    pars = ast.literal_eval(pars[0][0])
                    asy = pars['asy'][0]
                    con.close()
                    # Update asymmetry factor of 56Ni
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET shape = ?''', (
                        "{'name': 'VoigtAsy', 'sigma': 200, 'offset': 343.0, 'asy': " + str(asy) +
                        " , 'gamma': 20, 'offsetSlope': 0}",))
                    con.commit()
                    con.close()
                    # Fix asymmetry factor
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''UPDATE Lines SET fixShape = ?''', (
                        "{'asy': True, 'offsetSlope': True, 'sigma': [0,200], 'offset': False, 'gamma': [0,50]}",))
                    con.commit()
                    con.close()
                    print('The Asymmetry Factor of file', file58, ' and run', self.run, 'is', asy)
            # start value for offset
            spec = XMLImporter(path=os.path.join(self.filepath, file))
            offset = (spec.cts[0][0][0] + spec.cts[0][0][-1]) / 2
            self.adj_offset(offset)
            # start value for center
            index = list(spec.cts[0][0]).index(max(spec.cts[0][0]))  # center DAC voltage
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (file,))
            acc_v = cur.fetchall()[0][0]
            con.close()
            center_v = acc_v - spec.x[0][index]  # center total voltage
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT laserFreq FROM Files WHERE file = ? ''', (file,))
            laser_frequ = cur.fetchall()[0][0]
            con.close()
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT type FROM Files WHERE file = ?''', (file,))
            iso = cur.fetchall()[0][0]
            con.close()
            # set mass for doppler
            if iso == '58Ni':
                mass = 58
            elif iso == '60Ni':
                mass = 60
            else:
                mass = 56
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT frequency FROM Lines WHERE lineVar = ? ''', ('58_0',))
            frequ = cur.fetchall()[0][0]
            con.close()
            # calculate relative velocity for doppler
            v = Physics.relVelocity(Physics.qe * center_v, mass * Physics.u)
            v = -v
            # calculate doppler shifted frequency
            center_frequ = Physics.relDoppler(laser_frequ, v) - frequ
            print('Dopplershifted Frequ:', center_frequ)
            center = center_frequ - 500  # no idea why...
            self.adj_center(center)

            # Fit
            BatchFit.batchFit(np.array([file]), self.db, self.run)

            # Set asymmetry factor free again
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''UPDATE Lines SET fixShape = ?''', (
                "{'asy': [0,20], 'offsetSlope': True, 'sigma': [0,200], 'offset': False, 'gamma': [0,50]}",))
            con.commit()
            con.close()

    def single_cali(self):
        files = []
        shifts = []
        unc = []
        acc_volt = 29850
        for tup in self.calTuples:
            t_shifts = []
            t_unc = []
            files.append(tup[0])
            print('Calibrating files', tup)
            file58 = 'BECOLA_' + str(tup[0]) + '.xml'
            file60 = 'BECOLA_' + str(tup[1]) + '.xml'
            for i, run in enumerate(self.runs):
                self.run = run
                self.lineVar = self.linVars[i]
                ishift, err = self.cal_ishift(tup)
                t_shifts.append(ishift)
                t_unc.append(err)

                # Get acceleration voltage
                con = sqlite3.connect(self.db)
                cur = con.cursor()
                cur.execute('''SELECT accVolt FROM Files WHERE file = ?''', (file58,))
                acc_volt = cur.fetchall()[0][0]
                con.close()

                # Save to excel
                self.lastRow = self.wsCalAll.max_row
                self.wsCalAll.cell(row=self.lastRow + 1, column=1, value=str(tup))
                self.wsCalAll.cell(row=self.lastRow + 1, column=2, value=acc_volt)
                self.wsCalAll.cell(row=self.lastRow + 1, column=3, value=ishift)

            # Calculate mean of all shifts
            weights = []
            for u in t_unc:
                weights.append(1 / u ** 2)
            waverage = np.average(t_shifts, weights=weights)
            sigma = np.std(t_shifts)
            shifts.append(waverage)
            unc.append(sigma)

            # Calculate differential Doppler shift
            diff_dopp58 = Physics.diffDoppler(850343799, acc_volt, 58)
            diff_dopp60 = Physics.diffDoppler(850343799, acc_volt, 60)

            # Calculate calibration
            cal_volt = acc_volt + (waverage - 509.1) / (diff_dopp60 - diff_dopp58)
            print('Uncalibrated voltage is:', acc_volt)
            #cal_volt = acc_volt + (waverage - 508.2) / (6)
            print('Calibrated acceleration voltage is', cal_volt)

            # Update acceleration voltage
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''UPDATE Files SET accVolt = ? WHERE file LIKE ? OR file LIKE ?''',
                        (cal_volt, file58, file60))
            con.commit()
            con.close()

            self.wsCalAll.cell(row=self.lastRow + 1, column=4, value=cal_volt)
            self.wbAll.save(self.xlsFileAll)

        return shifts, files, unc

    def cal_ishift(self, file_tuple):
        # Get center of reference
        file58 = 'BECOLA_' + str(file_tuple[0]) + '.xml'
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        # Get corresponding isotope
        cur.execute('''SELECT type FROM Files WHERE file = ? ''', (file58,))
        iso_type58 = cur.fetchall()[0][0]
        # Query fit results for file and isotope combo
        cur.execute('''SELECT pars FROM FitRes WHERE file = ? AND iso = ? AND run = ?''',
                    (file58, iso_type58, self.run))
        pars58 = cur.fetchall()
        con.close()
        pars58dict = ast.literal_eval(pars58[0][0])
        center58 = pars58dict['center']

        # Get center of isotope of interest
        file60 = 'BECOLA_' + str(file_tuple[1]) + '.xml'
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        # Get corresponding isotope
        cur.execute(
            '''SELECT type FROM Files WHERE file = ? ''', (file60,))
        iso_type60 = cur.fetchall()[0][0]
        # Query fit results for file and isotope combo
        cur.execute('''SELECT pars FROM FitRes WHERE file = ? AND iso = ? AND run = ?''',
                    (file60, iso_type60, self.run))
        pars60 = cur.fetchall()
        con.close()
        pars60dict = ast.literal_eval(pars60[0][0])
        center60 = pars60dict['center']

        # Calculate isotope shift
        ishift = center60[0] - center58[0]
        print('Isotope shift of', file_tuple[1], 'is', ishift)

        # Calculate uncertainty
        uncert = np.sqrt(np.square(center58[1]) + np.square(center60[1]))

        return ishift, uncert

    def adj_offset(self, offset):
        # get parameter dictionary
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT shape FROM Lines WHERE lineVar LIKE ? ''', (self.lineVar,))
        shape = cur.fetchall()[0][0]
        shape_dict = ast.literal_eval(shape)
        # replace value of offset parameter
        shape_dict['offset'] = offset
        # Update database
        cur.execute('''UPDATE Lines SET shape = ? ''', (str(shape_dict),))
        con.commit()
        con.close()
        print('Adjusted Offset to ', shape_dict['offset'])

    def adj_center(self, center):
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET center = ?''', (center,))
        con.commit()
        con.close()
        print('Adjusted Center to', center)

    def calc56ni(self):
        shifts = []
        uncert = []
        files = []
        r = 2
        for tup in self.calGroups56:
            print('Calculating isotope shift:')
            t_shifts = []
            t_uncert = []
            files.append(tup[1])
            # Each pmt is evaluated individually and then averaged
            for i, run in enumerate(self.runs):
                self.run = run
                self.lineVar = self.linVars[i]
                ishift, err = self.cal_ishift(tup)
                t_shifts.append(ishift)
                t_uncert.append(err)
                self.allRes.append(ishift)
                self.allUnc.append(err)
            # calculate weighted average
            weights = []
            for u in t_uncert:
                weights.append(1 / u ** 2)
            waverage = np.average(t_shifts, weights=weights)
            sigma = np.std(t_shifts)
            shifts.append(waverage)
            uncert.append(sigma)
            # Save to excel
            self.wsISAll.cell(row=r, column=1, value=str(tup[1]))
            self.wsISAll.cell(row=r, column=2, value=waverage)
            self.wsISAll.cell(row=r, column=3, value=sigma)
            r += 1
        self.wbAll.save(self.xlsFileAll)
        self.shifts56 = shifts
        self.uncert56 = uncert

    def average(self):
        wts = []
        for i in self.uncert56:
            item = 1 / (i ** 2)
            wts.append(item)
        waverage = np.average(self.shifts56, weights=wts)
        print('Weighted average: ' + str(waverage))

        last_row = self.wsISAll.max_row
        self.wsISAll.cell(row=last_row + 2, column=1, value='Weighted Mean')
        self.wsISAll.cell(row=last_row + 2, column=2, value=waverage)

        sigma = np.std(self.shifts56)
        print('Standard deviation: ' + str(sigma))

        self.wsISAll.cell(row=last_row + 2, column=3, value=sigma)
        self.wbAll.save(self.xlsFileAll)
        n = len(self.shifts56)

        standard_error = sigma / (n ** (1 / 2))
        print('Standard error of the mean: ' + str(standard_error))

        w_error = 1 / np.sqrt((sum(wts)))
        print('Weighted error:', w_error)

        measurements = list(range(1, n + 1))

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - standard_error, waverage + standard_error, alpha=0.2, linewidth=0,
                         color='g')
        plt.fill_between([0, n + 1], waverage - standard_error - self.systUncert,
                         waverage + standard_error + self.systUncert,
                         alpha=0.1, linewidth=0, color='r')
        plt.title('Standard error')
        plt.show()

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - sigma, waverage + sigma, alpha=0.2, linewidth=0, color='b')
        plt.fill_between([0, n + 1], waverage - sigma - self.systUncert, waverage + sigma + self.systUncert, alpha=0.2,
                         linewidth=0,
                         color='r')
        plt.title('Standard deviation')
        plt.show()

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - w_error, waverage + w_error, alpha=0.2, linewidth=0, color='y')
        plt.fill_between([0, n + 1], waverage - w_error - self.systUncert, waverage + w_error + self.systUncert,
                         alpha=0.2, linewidth=0,
                         color='r')
        plt.title('Weighted error')
        plt.show()

        plt.errorbar(measurements, self.shifts56, yerr=self.uncert56, fmt='o')
        plt.plot([0, n + 1], [waverage, waverage], 'r-', label="mean isotope shift")
        plt.fill_between([0, n + 1], waverage - sigma / np.sqrt(n - 1), waverage + sigma / np.sqrt(n - 1), alpha=0.2,
                         linewidth=0, color='b')
        plt.fill_between([0, n + 1], waverage - sigma / np.sqrt(n - 1) - self.systUncert,
                         waverage + sigma / np.sqrt(n - 1) + self.systUncert, alpha=0.2, linewidth=0, color='r')
        plt.title('Standard deviation / sqrt(N-1)')
        plt.show()

    def plot_all(self):
        meas = list(range(0,len(self.allRes)))
        weights = []
        for u in self.allUnc:
            weights.append(1 / u ** 2)
        waverage = np.average(self.allRes, weights=weights)
        sigma = np.std(self.allRes)
        plt.errorbar(meas, self.allRes, yerr=self.allUnc, fmt='bo')
        plt.plot([0,len(self.allRes)], [waverage,waverage], 'r')
        plt.fill_between([0, len(self.allRes)], waverage - sigma, waverage + sigma, alpha=0.2, linewidth=0,
                         color='b')
        plt.fill_between([0, len(self.allRes)], waverage - sigma - self.systUncert,
                         waverage + sigma + self.systUncert,
                         alpha=0.1, linewidth=0, color='r')
        plt.show()
        print('Isotope shift averaged over all PMTs is', waverage)
        print('Standard deviation is', sigma)

    def stack_files(self, file_list):
        track = 0
        scalers = [0, 1, 2]

        # prepare arrays
        bin = 3
        voltage = [np.arange(-261, -42, bin), np.arange(-261, -42, bin), np.arange(-261, -42, bin)] # one list for each scaler
        sumcts = [np.zeros(len(voltage[0])), np.zeros(len(voltage[0])), np.zeros(len(voltage[0]))]
        sumbg = [np.zeros(len(voltage[0])), np.zeros(len(voltage[0])), np.zeros(len(voltage[0]))]
        scans = [np.zeros(len(voltage[0])), np.zeros(len(voltage[0])), np.zeros(len(voltage[0]))]
        err = [[], [], []]

        # iterate through scalers
        for s in scalers:
            print('scaler ', s)
            t0, t_width = self.find_timegates(file_list, track, s)
            t_min = (t0 - 2 * t_width) / 100
            t_max = (t0 + 2 * t_width) / 100
            # iterate through files and sum up
            for f in file_list:
                # spectrum only in the specified time gate
                spec = XMLImporter(path=self.workingDir + '\\data\\' + str(f),
                                   softw_gates=[[-350, 0, t_min, t_max], [-350, 0, t_min, t_max],
                                                [-350, 0, t_min, t_max]])
                # spectrum of background
                off = 200
                bg = XMLImporter(path=self.workingDir + '\\data\\' + str(f),
                                 softw_gates=[[-350, 0, t_min + off, t_max + off], [-350, 0, t_min + off, t_max + off],
                                              [-350, 0, t_min + off, t_max + off]])
                volcts = []
                for j, x in enumerate(spec.x[track]):
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''SELECT accVolt from Files WHERE file  = ?''', (f,))
                    accV = cur.fetchall()[0][0]
                    con.close()
                    offset = accV -29850
                    volcts.append((x + offset, spec.cts[track][s][j]))
                    for i, v in enumerate(voltage[s]):
                        if v - bin / 2 < x <= v + bin / 2:
                            sumcts[s][i] += spec.cts[track][s][j]
                            sumbg[s][i] += bg.cts[track][s][j]
                            scans[s][i] += spec.nrScans[track]
                plt.plot(spec.x[track], spec.cts[track][s])
                plt.plot(bg.x[track], bg.cts[track][s])
            plt.title('Scaler ' + str(s))
            plt.show()
            zeroInd = np.where(sumbg[s] == 0)
            sumcts[s] = np.delete(sumcts[s], zeroInd)
            sumbg[s] = np.delete(sumbg[s], zeroInd)
            voltage[s] = np.delete(voltage[s], zeroInd)
            scans[s] = np.delete(scans[s], zeroInd)
            plt.plot(voltage[s], sumcts[s], 'r.')
            plt.title('Scaler' + str(s))
            plt.show()
            plt.plot(voltage[s], sumcts[s] / scans[s], 'b.')
            plt.title('Scaler' + str(s))
            plt.show()
            # Calculate uncertainties for each cts value
            for cts in sumcts[s]:
                err[s].append(np.sqrt(cts))

        # Split data into  3 different tracks
        v0 = [[], [], []]
        cts = [[[], [], []], [[], [], []], [[], [], []]]
        unc = [[[], [], []], [[], [], []], [[], [], []]]
        nr_scans = [[[], [], []], [[], [], []], [[], [], []]]
        for s in scalers:
            for i, v in enumerate(voltage[s]):
                if v < -155:
                    if v not in v0[0]:
                        v0[0].append(v)
                    cts[0][s].append(sumcts[s][i])
                    unc[0][s].append(err[s][i])
                    nr_scans[0][s].append(scans[s][i])
                elif -155 < v < -75:
                    if v not in v0[1]:
                        v0[1].append(v)
                    cts[1][s].append(sumcts[s][i])
                    unc[1][s].append(err[s][i])
                    nr_scans[1][s].append(scans[s][i])
                else:
                    if v not in v0[2]:
                        v0[2].append(v)
                    cts[2][s].append(sumcts[s][i])
                    unc[2][s].append(err[s][i])
                    nr_scans[2][s].append(scans[s][i])
        for i, v in enumerate(v0):
            for s in scalers:
                plt.errorbar(v, cts[i][s], yerr=unc[i][s], fmt='b.')
                plt.title('track ' + str(i) + ', scaler ' + str(s))
                plt.show()

        for t in [0, 1, 2]:
            for s in scalers:
                for i, c in enumerate(cts[t][s]):
                    cts[t][s][i] = int(c)
        for t in [0, 1, 2]:
            for s in scalers:
                sum_n = 0
                for n in nr_scans[t][s]:
                    sum_n += n
                nr_scans[t][s] = sum_n // len(nr_scans[t][s])

        scaler_array = [[], [], []]
        for t in [0, 1, 2]:
            for s in scalers:
                timestep = 0
                for i, c in enumerate(cts[t][s]):
                    scaler_array[t].append((s, i, timestep, int(c)))
                    timestep += 1


        # Create dictionary for xml export
        file_creation_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        header_dict = {'type': 'trs',
                       'isotope': '55Ni',
                       'isotopeStartTime': file_creation_time,
                       'accVolt': 29850,
                       'laserFreq': Physics.wavenumber(self.laserFreq55) / 2,
                       'nOfTracks': 3,
                       'version': 99.0}
        track0_dict_header = {'trigger': {},  # Need a trigger dict!
                             'activePmtList': [0, 1, 2],  # Must be in form [0,1,2]
                             'colDirTrue': True,
                             'dacStartRegister18Bit': 0,
                             'dacStartVoltage': v0[0][0],
                             'dacStepSize18Bit': None,  # old format xml importer checks whether val or None
                             'dacStepsizeVoltage': bin,
                             'dacStopRegister18Bit': len(v0[0])-1,  # not real but should do the trick
                             'dacStopVoltage': v0[0][-1],
                             'invertScan': False,
                             'nOfBins': len(voltage[0]),
                             'nOfCompletedSteps': float(len(v0[0])),
                             'nOfScans': nr_scans[0][0],
                             'nOfSteps': len((v0[0])),
                             'postAccOffsetVolt': 0,
                             'postAccOffsetVoltControl': 0,
                             'softwGates': [[-261, -156, 0, 0.4], [-261, -156, 0, 0.4], [-261, -156, 0, 0.4]],
                             # For each Scaler: [DAC_Start_Volt, DAC_Stop_Volt, scaler_delay, softw_Gate_width]
                             'workingTime': [file_creation_time, file_creation_time],
                             'waitAfterReset1us': 0,  # looks like I need those for the importer
                             'waitForKepco1us': 0  # looks like I need this too
                             }
        track1_dict_header = {'trigger': {},  # Need a trigger dict!
                              'activePmtList': [0, 1, 2],  # Must be in form [0,1,2]
                              'colDirTrue': True,
                              'dacStartRegister18Bit': 0,
                              'dacStartVoltage': v0[1][0],
                              'dacStepSize18Bit': None,  # old format xml importer checks whether val or None
                              'dacStepsizeVoltage': bin,
                              'dacStopRegister18Bit': len(v0[1]) - 1,  # not real but should do the trick
                              'dacStopVoltage': v0[1][-1],
                              'invertScan': False,
                              'nOfBins': len(voltage[1]),
                              'nOfCompletedSteps': float(len(v0[1])),
                              'nOfScans': nr_scans[1][0],
                              'nOfSteps': len((v0[1])),
                              'postAccOffsetVolt': 0,
                              'postAccOffsetVoltControl': 0,
                              'softwGates': [[-261, -156, 0, 0.4], [-261, -156, 0, 0.4], [-261, -156, 0, 0.4]],
                              # For each Scaler: [DAC_Start_Volt, DAC_Stop_Volt, scaler_delay, softw_Gate_width]
                              'workingTime': [file_creation_time, file_creation_time],
                              'waitAfterReset1us': 0,  # looks like I need those for the importer
                              'waitForKepco1us': 0  # looks like I need this too
                              }
        track2_dict_header = {'trigger': {},  # Need a trigger dict!
                              'activePmtList': [0, 1, 2],  # Must be in form [0,1,2]
                              'colDirTrue': True,
                              'dacStartRegister18Bit': 0,
                              'dacStartVoltage': v0[2][0],
                              'dacStepSize18Bit': None,  # old format xml importer checks whether val or None
                              'dacStepsizeVoltage': bin,
                              'dacStopRegister18Bit': len(v0[2]) - 1,  # not real but should do the trick
                              'dacStopVoltage': v0[2][-1],
                              'invertScan': False,
                              'nOfBins': len(voltage[2]),
                              'nOfCompletedSteps': float(len(v0[2])),
                              'nOfScans': nr_scans[2][0],
                              'nOfSteps': len((v0[2])),
                              'postAccOffsetVolt': 0,
                              'postAccOffsetVoltControl': 0,
                              'softwGates': [[-261, -156, 0, 0.4], [-261, -156, 0, 0.4], [-261, -156, 0, 0.4]],
                              # For each Scaler: [DAC_Start_Volt, DAC_Stop_Volt, scaler_delay, softw_Gate_width]
                              'workingTime': [file_creation_time, file_creation_time],
                              'waitAfterReset1us': 0,  # looks like I need those for the importer
                              'waitForKepco1us': 0  # looks like I need this too
                              }

        data = '['
        for i, s in enumerate(scaler_array[0]):
            data = data + str(scaler_array[0][i]) + ' '
        data = data[:len(data)-1]
        data = data + ']'
        track0_dict_data = {
            'scalerArray_explanation': 'continously acquired data. List of Lists, each list represents the counts of '
                                       'one scaler as listed in activePmtList.Dimensions are: (len(activePmtList), '
                                       'nOfSteps), datatype: np.int32',
            'scalerArray': data}

        #if unc[0][0] is not None:
            #track0_dict_data['errorArray'] = unc[0]
            #track0_dict_data['errorArray_explanation'] = 'Optional: Non-standard errors. If this was not present, ' \
                                                        #'np.sqrt() would be used for errors during XML import. ' \
                                                        #'List of lists, each list represents the errors of one scaler '\
                                                        #'as listed in activePmtList.Dimensions  are: ' \
                                                        #'(len(activePmtList), nOfSteps), datatype: np.int32'

        data = '['
        for i, s in enumerate(scaler_array[1]):
            data = data + str(scaler_array[1][i]) + ' '
        data = data[:len(data) - 1]
        data = data + ']'
        track1_dict_data = {
            'scalerArray_explanation': 'continously acquired data. List of Lists, each list represents the counts of '
                                       'one scaler as listed in activePmtList.Dimensions are: (len(activePmtList), '
                                       'nOfSteps), datatype: np.int32',
            'scalerArray': data}

        #if unc[1][0] is not None:
            #track1_dict_data['errorArray'] = unc[1]
            #track1_dict_data['errorArray_explanation'] = 'Optional: Non-standard errors. If this was not present, ' \
                                                        #'np.sqrt() would be used for errors during XML import. ' \
                                                        #'List of lists, each list represents the errors of one scaler '\
                                                        #'as listed in activePmtList.Dimensions  are: ' \
                                                        #'(len(activePmtList), nOfSteps), datatype: np.int32'
        data = '['
        for i, s in enumerate(scaler_array[2]):
            data = data + str(scaler_array[2][i]) + ' '
        data = data[:len(data) - 1]
        data = data + ']'
        track2_dict_data = {
            'scalerArray_explanation': 'continously acquired data. List of Lists, each list represents the counts of '
                                       'one scaler as listed in activePmtList.Dimensions are: (len(activePmtList), '
                                       'nOfSteps), datatype: np.int32',
            'scalerArray': data}

        #if unc[2][0] is not None:
            #track2_dict_data['errorArray'] = unc[2]
            #track2_dict_data['errorArray_explanation'] = 'Optional: Non-standard errors. If this was not present, ' \
                                                        #'np.sqrt() would be used for errors during XML import. ' \
                                                        #'List of lists, each list represents the errors of one scaler '\
                                                        #'as listed in activePmtList.Dimensions  are: ' \
                                                        #'(len(activePmtList), nOfSteps), datatype: np.int32'

        v_proj = []
        for s in scalers:
            for v in cts[0][s]:
                v_proj.append(v)

        track0_vol_proj = {'voltage_projection': np.array(cts[0]),
                           'voltage_projection_explanation': 'voltage_projection of the time resolved data. List of '
                                                             'Lists, each list represents the counts of one scaler as '
                                                             'listed in activePmtList.Dimensions are: '
                                                             '(len(activePmtList), nOfSteps), datatype: np.int32'}

        v_proj = []
        for s in scalers:
            for v in cts[1][s]:
                v_proj.append(v)
        track1_vol_proj = {'voltage_projection': np.array(cts[1]),
                           'voltage_projection_explanation': 'voltage_projection of the time resolved data. List of '
                                                             'Lists, each list represents the counts of one scaler as '
                                                             'listed in activePmtList.Dimensions are: '
                                                             '(len(activePmtList), nOfSteps), datatype: np.int32'}
        v_proj = []
        for s in scalers:
            for v in cts[2][s]:
                v_proj.append(v)
        track2_vol_proj = {'voltage_projection': np.array(cts[2]),
                           'voltage_projection_explanation': 'voltage_projection of the time resolved data. List of '
                                                             'Lists, each list represents the counts of one scaler as '
                                                             'listed in activePmtList.Dimensions are: '
                                                             '(len(activePmtList), nOfSteps), datatype: np.int32'}

        dictionary = {'header': header_dict,
                      'tracks': {'track0': {'header': track0_dict_header,
                                            'data': track0_dict_data,
                                            'projections': track0_vol_proj
                                            },
                                 'track1': {'header': track1_dict_header,
                                            'data': track1_dict_data,
                                            'projections': track1_vol_proj
                                            },
                                 'track2': {'header': track2_dict_header,
                                            'data': track2_dict_data,
                                            'projections': track2_vol_proj
                                            },
                                 }
                      }

        root = ET.Element('BecolaData')

        xmlWriteDict(root, dictionary)
        xml = ET.ElementTree(root)
        xml.write(self.workingDir + '\\data\\BECOLA_Stacked.xml')

        # Add to database
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''INSERT OR IGNORE INTO Files (file, filePath, date, type) VALUES (?, ?, ?, ?)''',
                    ('BECOLA_Stacked.xml', 'data\BECOLA_Stacked.xml', file_creation_time, '55Ni' + '_sum'))
        con.commit()
        cur.execute(
            '''UPDATE Files SET offset = ?, accVolt = ?,  laserFreq = ?, laserFreq_d = ?, colDirTrue = ?, 
            voltDivRatio = ?, lineMult = ?, lineOffset = ?, errDateInS = ? WHERE file = ? ''',
            ('[0]', 29850, self.laserFreq55, 0, True, str({'accVolt': 1.0, 'offset': 1.0}), 1, 0,
             1, 'BECOLA_Stacked.xml'))
        con.commit()
        con.close()

        BatchFit.batchFit(['BECOLA_Stacked.xml'], self.db_stacked, 'Voigt55', x_as_voltage=True,
                          save_file_as='.png')

        #BatchFit.batchFit(['BECOLA_Stacked.xml'], self.db_stacked, 'AsymVoigt55', x_as_voltage=True,
                          #save_file_as='.png')

    def stack_files_calibrated(self, file_list):
        track = 0
        scalers = [0, 1, 2]

        # prepare arrays
        bin = 3
        voltage = [np.arange(-261, -42, bin), np.arange(-261, -42, bin), np.arange(-261, -42, bin)] # one list for each scaler
        sumcts = [np.zeros(len(voltage[0])), np.zeros(len(voltage[0])), np.zeros(len(voltage[0]))]
        sumbg = [np.zeros(len(voltage[0])), np.zeros(len(voltage[0])), np.zeros(len(voltage[0]))]
        scans = [np.zeros(len(voltage[0])), np.zeros(len(voltage[0])), np.zeros(len(voltage[0]))]
        err = [[], [], []]

        # iterate through scalers
        for s in scalers:
            print('scaler ', s)
            t0, t_width = self.find_timegates(file_list, track, s)
            t_min = (t0 - 2 * t_width) / 100
            t_max = (t0 + 2 * t_width) / 100
            # iterate through files and sum up
            volcts = []
            for f in file_list:
                # spectrum only in the specified time gate
                spec = XMLImporter(path=self.workingDir + '\\data\\' + str(f),
                                   softw_gates=[[-350, 0, t_min, t_max], [-350, 0, t_min, t_max],
                                                [-350, 0, t_min, t_max]])
                # spectrum of background
                off = 200
                bg = XMLImporter(path=self.workingDir + '\\data\\' + str(f),
                                 softw_gates=[[-350, 0, t_min + off, t_max + off], [-350, 0, t_min + off, t_max + off],
                                              [-350, 0, t_min + off, t_max + off]])
                for j, x in enumerate(spec.x[track]):
                    con = sqlite3.connect(self.db)
                    cur = con.cursor()
                    cur.execute('''SELECT accVolt from Files WHERE file  = ?''', (f,))
                    accV = cur.fetchall()[0][0]
                    con.close()
                    offset = accV - 29850
                    volcts.append((x - offset, spec.cts[track][s][j], spec.nrScans[track], bg.cts[track][s][j]))
                #print(max(spec.x[track]) - offset)
                #if max(spec.x[track]) - offset > 0:
                    #print(f)
                plt.plot(spec.x[track], spec.cts[track][s])
                plt.plot(bg.x[track], bg.cts[track][s])
            plt.xlim(-300, 60)
            plt.title('Uncalibrated, Scaler ' + str(s))
            plt.show()
            v = np.arange(-260, -33, bin)
            sumc = np.zeros(len(v))
            sumb = np.zeros(len(v))
            sc = np.zeros(len(v))
            for tup in volcts:
                for j, item in enumerate(v):
                    if item - bin / 2 < tup[0] <= item + bin / 2:
                        sumc[j] += tup[1]
                        sc[j] += tup[2]
                        sumb[j] += tup[3]
            zInd = np.where(sumb == 0)
            sumc = np.delete(sumc, zInd)
            sumb = np.delete(sumb, zInd)
            v = np.delete(v, zInd)
            sc = np.delete(sc, zInd)
            #print(sumc)
            plt.plot(v, sumc, 'b.')
            plt.title('Calibrated and summed, Sclaer' + str(s))
            plt.show()

            unc = []
            for cts in sumc:
                unc.append(np.sqrt(cts))
            voltage[s] = v
            sumcts[s] = sumc
            sumbg[s] = sumb
            scans[s] = sc
            err[s] = unc

            #plt.title('Scaler' + str(s))
            #plt.show()

        print('Sumcts:', sumcts)
        # Split data into  3 different tracks
        v0 = [[[], [], []], [[], [], []], [[], [], []]]
        cts = [[[], [], []], [[], [], []], [[], [], []]]
        unc = [[[], [], []], [[], [], []], [[], [], []]]
        nr_scans = [[[], [], []], [[], [], []], [[], [], []]]
        for s in scalers:
            print(len(sumcts[s]))
            print(len(voltage[s]))
            for i, v in enumerate(voltage[s]):
                if v <= -177:
                    if v not in v0[0][s]:
                        v0[0][s].append(v)
                    cts[0][s].append(sumcts[s][i])
                    unc[0][s].append(err[s][i])
                    nr_scans[0][s].append(scans[s][i])
                elif -177 < v <= -110:
                    if v not in v0[1][s]:
                        v0[1][s].append(v)
                    cts[1][s].append(sumcts[s][i])
                    unc[1][s].append(err[s][i])
                    nr_scans[1][s].append(scans[s][i])
                else:
                    if v not in v0[2][s]:
                        v0[2][s].append(v)
                    cts[2][s].append(sumcts[s][i])
                    unc[2][s].append(err[s][i])
                    nr_scans[2][s].append(scans[s][i])

        for s in scalers:
            for track in [0, 1, 2]:
                plt.errorbar(v0[track][s], cts[track][s], yerr=unc[track][s], fmt='b.')
                plt.title('track ' + str(track) + ', scaler ' + str(s))
                plt.show()

        for t in [0, 1, 2]:
            for s in scalers:
                for i, c in enumerate(cts[t][s]):
                    cts[t][s][i] = int(c)
        for t in [0, 1, 2]:
            for s in scalers:
                sum_n = 0
                for n in nr_scans[t][s]:
                    sum_n += n
                nr_scans[t][s] = sum_n // len(nr_scans[t][s])

        # Make all voltage lists the same length
        for track in [0, 1, 2]:
            #print(v0[track][0], v0[track][1], v0[track][2])
            #print(v0[track][0] == v0[track][1])
            while v0[track][0] != v0[track][1]:
                #print(len(v0[track][0]), len ( v0[track][1]))
                if len(v0[track][0]) < len(v0[track][1]):
                    for i, v in enumerate(v0[track][0]):
                        #print('len0 < len1')
                        #print('v0:', v0[track][0])
                        #print('v1:', v0[track][1])
                        if v < v0[track][1][i]:
                            #print('add', v)
                            v0[track][1].insert(i, v)
                            cts[track][1].insert(i, 0)
                            #print('v1:', v0[track][1])
                            break
                        elif v > v0[track][1][i]:
                            #print('add', v0[track][1][i])
                            v0[track][0].insert(i, v0[track][1][i])
                            cts[track][0].insert(i, 0)
                            #print('v0:', v0[track][0])
                            break
                if len(v0[track][0]) > len(v0[track][1]):
                    for i, v in enumerate(v0[track][1]):
                        #print('len0 > len1')
                        #print('v0:', v0[track][0])
                        #print('v1:', v0[track][1])
                        #print('v1:', v, 'v0:', v0[track][0][i])
                        if v < v0[track][0][i]:
                            #print('add', v)
                            v0[track][0].insert(i, v)
                            cts[track][0].insert(i, 0)
                            #print('v1:', v0[track][0])
                            break
                        elif v > v0[track][0][i]:
                            #print('add', v0[track][0][i])
                            v0[track][1].insert(i, v0[track][0][i])
                            cts[track][1].insert(i, 0)
                            #print('v0:', v0[track][0])
                            break
                        elif i == len(v0[track][1])-1:
                            #print('add', v0[track][0][i+1])
                            v0[track][1].append(v0[track][0][i+1])
                            cts[track][1].append(0)
                            #print('v1:', v0[track][1])
                            break
            while v0[track][0] != v0[track][2]:
                if len(v0[track][0]) > len(v0[track][2]):
                    for i, v in enumerate(v0[track][0]):
                        #print('len0 > len2')
                        #print('v0:', v0[track][0])
                        #print('v2:', v0[track][2])
                        if v < v0[track][2][i]:
                            #print('add', v)
                            v0[track][2].insert(i, v)
                            cts[track][2].insert(i, 0)
                            #print('v2:', v0[track][2])
                            break
                        elif v > v0[track][2][i]:
                            #print('add', v0[track][2][i])
                            v0[track][0].insert(i, v0[track][2][i])
                            cts[track][0].insert(i, 0)
                            #print('v0:', v0[track][0])
                            break
                elif len(v0[track][0]) < len(v0[track][2]):
                    for i, v in enumerate(v0[track][2]):
                        #print('len0 < len2')
                        #print('v0:', v0[track][0])
                        #print('v2:', v0[track][2])
                        if i < len(v0[track][1]):
                            if v < v0[track][0][i]:
                                #print('add', v)
                                v0[track][0].insert(i, v)
                                cts[track][0].insert(i, 0)
                                #print('v2:', v0[track][0])
                                break
                            elif v > v0[track][0][i]:
                                #print('add', v0[track][0][i])
                                v0[track][2].insert(i, v0[track][0][i])
                                cts[track][2].insert(i, 0)
                                #print('v0:', v0[track][2])
                                break
                        elif i >= len(v0[track][1]):
                            #print('add', v)
                            v0[track][0].insert(i, v)
                            cts[track][0].insert(i, 0)
                            #print('v0:', v0[track][0])
                            break

            if v0[track][0] != v0[track][2]:
                for i, v in enumerate(v0[track][0]):
                    if v != v0[track][2][i]:
                        print('add', v)

        scaler_array = [[], [], []]
        for t in [0, 1, 2]:
            for s in scalers:
                timestep = 0
                for i, c in enumerate(cts[t][s]):
                    scaler_array[t].append((s, i, timestep, int(c)))
                    timestep += 1


        # Create dictionary for xml export
        file_creation_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        header_dict = {'type': 'trs',
                       'isotope': '55Ni',
                       'isotopeStartTime': file_creation_time,
                       'accVolt': 29850,
                       'laserFreq': Physics.wavenumber(self.laserFreq55) / 2,
                       'nOfTracks': 3,
                       'version': 99.0}
        track0_dict_header = {'trigger': {},  # Need a trigger dict!
                             'activePmtList': [0, 1, 2],  # Must be in form [0,1,2]
                             'colDirTrue': True,
                             'dacStartRegister18Bit': 0,
                             'dacStartVoltage': v0[0][0][0],
                             'dacStepSize18Bit': None,  # old format xml importer checks whether val or None
                             'dacStepsizeVoltage': bin,
                             'dacStopRegister18Bit': len(v0[0][0])-1,  # not real but should do the trick
                             'dacStopVoltage': v0[0][0][-1],
                             'invertScan': False,
                             'nOfBins': len(voltage[0]),
                             'nOfCompletedSteps': float(len(v0[0][0])),
                             'nOfScans': nr_scans[0][0],
                             'nOfSteps': len((v0[0][0])),
                             'postAccOffsetVolt': 0,
                             'postAccOffsetVoltControl': 0,
                             'softwGates': [[-261, -156, 0, 0.4], [-261, -156, 0, 0.4], [-261, -156, 0, 0.4]],
                             # For each Scaler: [DAC_Start_Volt, DAC_Stop_Volt, scaler_delay, softw_Gate_width]
                             'workingTime': [file_creation_time, file_creation_time],
                             'waitAfterReset1us': 0,  # looks like I need those for the importer
                             'waitForKepco1us': 0  # looks like I need this too
                             }
        track1_dict_header = {'trigger': {},  # Need a trigger dict!
                              'activePmtList': [0, 1, 2],  # Must be in form [0,1,2]
                              'colDirTrue': True,
                              'dacStartRegister18Bit': 0,
                              'dacStartVoltage': v0[1][0][0],
                              'dacStepSize18Bit': None,  # old format xml importer checks whether val or None
                              'dacStepsizeVoltage': bin,
                              'dacStopRegister18Bit': len(v0[1][0]) - 1,  # not real but should do the trick
                              'dacStopVoltage': v0[1][0][-1],
                              'invertScan': False,
                              'nOfBins': len(voltage[1]),
                              'nOfCompletedSteps': float(len(v0[1][0])),
                              'nOfScans': nr_scans[1][0],
                              'nOfSteps': len((v0[1][0])),
                              'postAccOffsetVolt': 0,
                              'postAccOffsetVoltControl': 0,
                              'softwGates': [[-261, -156, 0, 0.4], [-261, -156, 0, 0.4], [-261, -156, 0, 0.4]],
                              # For each Scaler: [DAC_Start_Volt, DAC_Stop_Volt, scaler_delay, softw_Gate_width]
                              'workingTime': [file_creation_time, file_creation_time],
                              'waitAfterReset1us': 0,  # looks like I need those for the importer
                              'waitForKepco1us': 0  # looks like I need this too
                              }
        track2_dict_header = {'trigger': {},  # Need a trigger dict!
                              'activePmtList': [0, 1, 2],  # Must be in form [0,1,2]
                              'colDirTrue': True,
                              'dacStartRegister18Bit': 0,
                              'dacStartVoltage': v0[2][0][0],
                              'dacStepSize18Bit': None,  # old format xml importer checks whether val or None
                              'dacStepsizeVoltage': bin,
                              'dacStopRegister18Bit': len(v0[2][0]) - 1,  # not real but should do the trick
                              'dacStopVoltage': v0[2][0][-1],
                              'invertScan': False,
                              'nOfBins': len(voltage[2]),
                              'nOfCompletedSteps': float(len(v0[2][0])),
                              'nOfScans': nr_scans[2][0],
                              'nOfSteps': len((v0[2][0])),
                              'postAccOffsetVolt': 0,
                              'postAccOffsetVoltControl': 0,
                              'softwGates': [[-261, -156, 0, 0.4], [-261, -156, 0, 0.4], [-261, -156, 0, 0.4]],
                              # For each Scaler: [DAC_Start_Volt, DAC_Stop_Volt, scaler_delay, softw_Gate_width]
                              'workingTime': [file_creation_time, file_creation_time],
                              'waitAfterReset1us': 0,  # looks like I need those for the importer
                              'waitForKepco1us': 0  # looks like I need this too
                              }

        data = '['
        for i, s in enumerate(scaler_array[0]):
            data = data + str(scaler_array[0][i]) + ' '
        data = data[:len(data)-1]
        data = data + ']'
        track0_dict_data = {
            'scalerArray_explanation': 'continously acquired data. List of Lists, each list represents the counts of '
                                       'one scaler as listed in activePmtList.Dimensions are: (len(activePmtList), '
                                       'nOfSteps), datatype: np.int32',
            'scalerArray': data}

        #if unc[0][0] is not None:
            #track0_dict_data['errorArray'] = unc[0]
            #track0_dict_data['errorArray_explanation'] = 'Optional: Non-standard errors. If this was not present, ' \
                                                        #'np.sqrt() would be used for errors during XML import. ' \
                                                        #'List of lists, each list represents the errors of one scaler '\
                                                        #'as listed in activePmtList.Dimensions  are: ' \
                                                        #'(len(activePmtList), nOfSteps), datatype: np.int32'

        data = '['
        for i, s in enumerate(scaler_array[1]):
            data = data + str(scaler_array[1][i]) + ' '
        data = data[:len(data) - 1]
        data = data + ']'
        track1_dict_data = {
            'scalerArray_explanation': 'continously acquired data. List of Lists, each list represents the counts of '
                                       'one scaler as listed in activePmtList.Dimensions are: (len(activePmtList), '
                                       'nOfSteps), datatype: np.int32',
            'scalerArray': data}

        #if unc[1][0] is not None:
            #track1_dict_data['errorArray'] = unc[1]
            #track1_dict_data['errorArray_explanation'] = 'Optional: Non-standard errors. If this was not present, ' \
                                                        #'np.sqrt() would be used for errors during XML import. ' \
                                                        #'List of lists, each list represents the errors of one scaler '\
                                                        #'as listed in activePmtList.Dimensions  are: ' \
                                                        #'(len(activePmtList), nOfSteps), datatype: np.int32'
        data = '['
        for i, s in enumerate(scaler_array[2]):
            data = data + str(scaler_array[2][i]) + ' '
        data = data[:len(data) - 1]
        data = data + ']'
        track2_dict_data = {
            'scalerArray_explanation': 'continously acquired data. List of Lists, each list represents the counts of '
                                       'one scaler as listed in activePmtList.Dimensions are: (len(activePmtList), '
                                       'nOfSteps), datatype: np.int32',
            'scalerArray': data}

        #if unc[2][0] is not None:
            #track2_dict_data['errorArray'] = unc[2]
            #track2_dict_data['errorArray_explanation'] = 'Optional: Non-standard errors. If this was not present, ' \
                                                        #'np.sqrt() would be used for errors during XML import. ' \
                                                        #'List of lists, each list represents the errors of one scaler '\
                                                        #'as listed in activePmtList.Dimensions  are: ' \
                                                        #'(len(activePmtList), nOfSteps), datatype: np.int32'

        v_proj = []
        for s in scalers:
            for v in cts[0][s]:
                v_proj.append(v)

        track0_vol_proj = {'voltage_projection': np.array(cts[0]),
                           'voltage_projection_explanation': 'voltage_projection of the time resolved data. List of '
                                                             'Lists, each list represents the counts of one scaler as '
                                                             'listed in activePmtList.Dimensions are: '
                                                             '(len(activePmtList), nOfSteps), datatype: np.int32'}

        v_proj = []
        for s in scalers:
            for v in cts[1][s]:
                v_proj.append(v)
        track1_vol_proj = {'voltage_projection': np.array(cts[1]),
                           'voltage_projection_explanation': 'voltage_projection of the time resolved data. List of '
                                                             'Lists, each list represents the counts of one scaler as '
                                                             'listed in activePmtList.Dimensions are: '
                                                             '(len(activePmtList), nOfSteps), datatype: np.int32'}
        v_proj = []
        for s in scalers:
            for v in cts[2][s]:
                v_proj.append(v)
        track2_vol_proj = {'voltage_projection': np.array(cts[2]),
                           'voltage_projection_explanation': 'voltage_projection of the time resolved data. List of '
                                                             'Lists, each list represents the counts of one scaler as '
                                                             'listed in activePmtList.Dimensions are: '
                                                             '(len(activePmtList), nOfSteps), datatype: np.int32'}

        dictionary = {'header': header_dict,
                      'tracks': {'track0': {'header': track0_dict_header,
                                            'data': track0_dict_data,
                                            'projections': track0_vol_proj
                                            },
                                 'track1': {'header': track1_dict_header,
                                            'data': track1_dict_data,
                                            'projections': track1_vol_proj
                                            },
                                 'track2': {'header': track2_dict_header,
                                            'data': track2_dict_data,
                                            'projections': track2_vol_proj
                                            },
                                 }
                      }

        root = ET.Element('BecolaData')

        xmlWriteDict(root, dictionary)
        xml = ET.ElementTree(root)
        xml.write(self.workingDir + '\\data\\BECOLA_Stacked.xml')

        # Add to database
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''INSERT OR IGNORE INTO Files (file, filePath, date, type) VALUES (?, ?, ?, ?)''',
                    ('BECOLA_Stacked.xml', 'data\BECOLA_Stacked.xml', file_creation_time, '55Ni' + '_sum'))
        con.commit()
        cur.execute(
            '''UPDATE Files SET offset = ?, accVolt = ?,  laserFreq = ?, laserFreq_d = ?, colDirTrue = ?, 
            voltDivRatio = ?, lineMult = ?, lineOffset = ?, errDateInS = ? WHERE file = ? ''',
            ('[0]', 29850, self.laserFreq55, 0, True, str({'accVolt': 1.0, 'offset': 1.0}), 1, 0,
             1, 'BECOLA_Stacked.xml'))
        con.commit()
        con.close()

    def fit_stacked(self, sym=True):
        con = sqlite3.connect(self.db_stacked)
        cur = con.cursor()
        cur.execute('''SELECT fixShape from Lines WHERE lineVar = ? ''', ('55_Asy',))
        shape = cur.fetchall()
        shape_dict = ast.literal_eval(shape[0][0])
        con.close()
        if sym:
            shape_dict['asy'] = True
        else:
            shape_dict['asy'] = False
        print(shape_dict['asy'])
        con = sqlite3.connect(self.db_stacked)
        cur = con.cursor()
        cur.execute('''UPDATE Lines SET fixShape = ?''', (str(shape_dict),))
        con.commit()
        con.close()

        BatchFit.batchFit(['BECOLA_Stacked.xml'], self.db_stacked, 'AsymVoigt55', x_as_voltage=True,
                          save_file_as='.png')

        con = sqlite3.connect(self.db_stacked)
        cur = con.cursor()
        cur.execute('''SELECT pars From FitRes WHERE run = ?''', ('AsymVoigt55',))
        paras = cur.fetchall()
        con.close()
        para_dict = ast.literal_eval(paras[0][0])
        print('Al =', para_dict['Al'], '\nAu =', para_dict['Au'], '\nBl =', para_dict['Bl'], '\nBu =', para_dict['Bu'])
        al = para_dict['Al'][0]
        au = para_dict['Au'][0]
        print('A relation =', str(au / al))

    def find_timegates(self, file_list, track, scaler):
        # returns the center of the timegate and sigma
        # param:    file_list: list of files to sum up
        #           track: which track is used
        #           scaler: which scaler is used

        sum_t_proj = np.zeros(1024) # prepare the list of counts for each time step

        # iterate through files and sum up the counts for each time step
        for file in file_list:
            sum_t_proj = list(map(add, sum_t_proj, self.t_proj(file, track, scaler)))

        time = list(range(0, len(sum_t_proj)))  # list of time steps (for fitting and plotting)

        # Fit a gaussian function to the time projection
        a, sigma, center, offset = self.fit_time(time, sum_t_proj)

        # Plot counts and fit
        plt.plot(time, sum_t_proj, 'b.')
        plt.title('Scaler' + str(scaler))
        plt.plot(time, self.gauss(time, a, sigma, center, offset), 'r-')
        plt.axvline(center - 2 * sigma, color='y')
        plt.axvline(center + 2 * sigma, color='y')
        plt.show()
        return center, sigma

    def t_proj(self, file, track, scaler):
        # returns the time projection
        # param:    file: xml-file to get the data of
        #           track: which track is used
        #           scaler: which pmt is used
        self.spec = XMLImporter(path=self.workingDir + '\\data\\' + file)
        return self.spec.t_proj[track][scaler]

    def fit_time(self, time, cts):
        # fits a gaussian to the time projection
        # param:    time: list of time steps
        #           cts: list of counts

        # guess Start parameters and set amplitude and sigma positive
        start_par = np.array([max(cts), 10, time[cts.index(max(cts))], (time[0]+time[-1]) / 2])
        param_bounds = ([0, 0, -np.inf, -np.inf], [np.inf, np.inf, np.inf, np.inf])

        # Fitting
        a, sigma, center, offset = curve_fit(self.gauss, time, cts, start_par, bounds=param_bounds)[0]
        return a, sigma, center, offset

    def gauss(self, t, a, s , t0, o):
        # prams:    t: time
        #           a: cts
        #           s: sigma
        #           t0: mid of time
        #           o: offset
        return o + a / np.sqrt(2 * np.pi * s ** 2) * np.exp(-1 / 2 * np.square((t - t0) / s))

    def plotVolt(self):
        files = []
        voltage = []
        cnt = 0
        for tup in self.calTuples:
            file58 = 'BECOLA_' + str(tup[0]) + '.xml'
            print(file58)
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT accVolt FROM files WHERE file LIKE ?''', (file58,))
            volt = cur.fetchall()[0][0]
            con.close()
            voltage.append(volt)
            files.append(cnt)
            cnt += 1
        print(voltage)
        plt.plot(files, voltage, 'b.')
        plt.title('Voltage')
        plt.show()

    def calcQ(self):
        b_ref = -103.3  # 61Ni from Kaufmann
        q_ref = 0.163   # 61Ni from Kaufmann
        con = sqlite3.connect(self.db_stacked)
        cur = con.cursor()
        cur.execute('''SELECT pars from FitRes WHERE run = ?''', ('AsymVoigt55',))
        pars = cur.fetchall()
        con.close()
        par_dict = ast.literal_eval(pars[0][0])
        b = par_dict['Bl'][0]
        q = b * q_ref / b_ref
        delta_b = par_dict['Bl'][1]
        delta_q_ref = 0.003
        delta_b_ref = 1.7
        delta_q = np.sqrt((q_ref / b_ref * delta_b) ** 2 + (b / b_ref * delta_q_ref) ** 2 +
                          (b * q_ref / (b_ref ** 2) * delta_b_ref) ** 2)
        print('Electric quadrupole moment:', q, '+/-', delta_q, 'b')

    def calcMu(self):
        con = sqlite3.connect(self.db_stacked)
        cur = con.cursor()
        cur.execute('''SELECT pars from FitRes WHERE run = ?''', ('AsymVoigt55',))
        pars = cur.fetchall()
        con.close()
        para_dict = ast.literal_eval(pars[0][0])
        A = para_dict['Al'][0]
        mu_ref = 0 - 0.7486    #in nuclear moments from Kaufmann
        A_ref = 0 - 455.0      # from Kaufmann
        I = 3.5
        I_ref = 1.5
        mu = A * mu_ref / A_ref * I / I_ref
        delta_A = para_dict['Al'][1]
        delta_mu_ref = 0.0004
        delta_A_ref = 0.3
        delta_mu = np.sqrt((mu / A * delta_A) ** 2 + (mu / mu_ref * delta_mu_ref) ** 2 + (mu / A_ref * delta_A_ref) ** 2)
        print('Magnetic dipole moment:', mu, '+/-', delta_mu, 'nuclear moments')

    def get_ref_center(self):
        centers58 = []
        errors58 = []
        centers60 = []
        errors60 = []
        for tup in self.calTuples:
            file58 = 'BECOLA_' + str(tup[0]) + '.xml'
            file60 = 'BECOLA_' + str(tup[1]) + '.xml'
            con = sqlite3.connect(self.db)
            cur = con.cursor()
            cur.execute('''SELECT pars from FitRes WHERE file = ? ''', (file58,))
            pars = cur.fetchall()[0][0]
            fit_parameter = ast.literal_eval(pars)
            centers58.append(fit_parameter['center'][0])
            errors58.append(fit_parameter['center'][1])
            cur.execute('''SELECT pars from FitRes WHERE file = ? ''', (file60,))
            pars = cur.fetchall()[0][0]
            fit_parameter = ast.literal_eval(pars)
            centers60.append(fit_parameter['center'][0])
            errors60.append(fit_parameter['center'][1])
            con.close()
        plt.errorbar([1,2,3,4,5], centers58, yerr=errors58, fmt='b.')
        plt.errorbar([1, 2, 3, 4, 5], centers60, yerr=errors60, fmt='r.')
        plt.title('Reference centers')
        plt.show()
        weights = []
        for e in errors58:
            weights.append(1 / (e ** 2))
        center = np.average(centers58, weights=weights)
        sigma = np.std(centers58)
        print('Mean center is', center, '+/-', sigma)
        return center, sigma

    def get_center55(self):
        file = 'BECOLA_Stacked.xml'
        con = sqlite3.connect(self.db_stacked)
        cur = con.cursor()
        cur.execute('''SELECT pars from FitRes WHERE file = ?''', (file,))
        pars = cur.fetchall()[0][0]
        parameter = ast.literal_eval(pars)
        con.close()
        return parameter['center'][0], parameter['center'][1]

def Analyze55Ni():
    calibration_tuples = [(6362, 6363), (6395, 6396), (6417, 6419), (6462, 6463), (6467, 6466)]
    calibration_groups = [((6362, 6395), (6369, 6373, 6370, 6375, 6376, 6377, 6378, 6380, 6382, 6383, 6384, 6387, 6391, 6392, 6393)),
                      ((6395,6417), (6399, 6400, 6401, 6402, 6404, 6405, 6406, 6408, 6410, 6411, 6412)),
                      ((6417,6462), (6428, 6429, 6430, 6431, 6432, 6433, 6434, 6436, 6438, 6440, 6441, 6444, 6445, 6447, 6448)),
                      ((6462,6467), (6468, 6470, 6471, 6472, 6473, 6478, 6479, 6480, 6493))]
    syst_uncert = 0.4

    niAna = NiAnalysis(calibration_tuples, calibration_groups, syst_uncert)
    #niAna.reset()
    #niAna.prep()
    #niAna.assign_cal()
    #print('---------------Voltage updated')

    niAna.plotVolt()
    file_list55 = niAna.get_files('55Ni')
    niAna.stack_files_calibrated(file_list55)
    niAna.fit_stacked(sym=True)
    center58, sigma58 = niAna.get_ref_center()
    center55, sigma55 = niAna.get_center55()
    uncert = np.sqrt(np.square(sigma58) + np.square(sigma55))
    print('55Ni center is', center55)
    print('Isotope shift of 55Ni is', center55 - center58, '+/-', uncert)
    niAna.calcQ()
    niAna.calcMu()

def Analyze56Ni():
    # Calibrate Voltage
    calTuples = [(6191, 6192), (6207, 6208), (6242, 6243), (6253, 6254)]

    # calGroups56 = [(6191,6202),(6191,6203),(6191,6204),(6207,6211),(6242,6238),(6242,6239),
    # (6242,6240),(6253,6251),(6253,6252)]
    # systUncert = 2.2

    calGroups56 = [(6242, 6238), (6242, 6239), (6242, 6240), (6253, 6251), (6253, 6252)]
    systUncert = 0.4

    # run = 'AsymVoigt0'
    # line = '58_0'
    # file = 'Scaler0.xlsx'
    # run = 'AsymVoigt1'
    # line = '58_1'
    # file = 'Scaler1.xlsx'
    # run = 'AsymVoigt2'
    # line = '58_2'
    # file = 'Scaler2.xlsx'
    # run = 'AsymVoigtAll'
    # line = '58_All'
    # file = 'AllScalers.xlsx'

    niAna = NiAnalysis(calTuples, calGroups56, systUncert)
    for i, run in enumerate(niAna.runs):
        niAna.run = run
        niAna.lineVar = niAna.linVars[i]
        niAna.xlsFile = niAna.workingDir + '\\' + niAna.files[i]
        niAna.calibrate()
        niAna.assign_cal()

        # Fit Ni 56
        filelist56 = niAna.get_files('56Ni')
        niAna.fit_all(filelist56)
        niAna.calc56ni()
        niAna.average()

    meas = list(range(0, 20))
    print('meas:', len(meas))
    print('Res:', len(niAna.allRes))
    weights = []
    for u in niAna.allUnc:
        weights.append(1 / u ** 2)
    waverage = np.average(niAna.allRes, weights=weights)
    print('Weighted Average:', waverage)
    sigma = np.std(niAna.allRes)
    print('Standard deviation:', sigma)
    plt.errorbar(meas, niAna.allRes, yerr=niAna.allUnc, fmt='bo')
    plt.plot([0, 20], [waverage, waverage], 'r')
    plt.fill_between([0, 20], waverage - sigma, waverage + sigma, alpha=0.2, linewidth=0,
                     color='b')
    plt.fill_between([0, 20], waverage - sigma - niAna.systUncert, waverage + sigma + niAna.systUncert, alpha=0.2,
                     linewidth=0,
                     color='r')
    plt.show()

#Analyze56Ni()
Analyze55Ni()