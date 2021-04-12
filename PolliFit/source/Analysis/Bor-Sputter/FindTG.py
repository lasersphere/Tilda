import sqlite3
import os
from openpyxl import Workbook, load_workbook
import BatchFit
from SPFitter import SPFitter
from Measurement import MeasLoad
from DBIsotope import DBIsotope
from Spectra.FullSpec import FullSpec
import ast
import numpy as np


#### Find timeGates with best signal to noise ratio and set database to the determined time gate ####

class FindTG:

    def __init__(self):
        #workingdir = 'C:\\Users\\Laura Renth\\Desktop\\Daten\\Promotion\\Bor\\Sputter source\\2021-03-Data' #working dir IKP
        #self.workingdir = 'C:\\Users\\Laura Renth\\ownCloud\\User\\Laura\\KOALA\\2021-03-Data'  #working dir IKP Owncloud
        self.workingdir = 'D:\\ownCloud\\User\\Laura\\KOALA\\2021-03-Data'  # working dir hp Owncloud
        self.db = os.path.join(self.workingdir, 'B-_Auswertung.sqlite')

        # isotope and run to investigate
        self.isotope = '11B_D2'
        self.run = 'sym2'
        #self.isotope = '11B'
        #self.run = 'sym1'

        # set line to fit
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT lineVar FROM Runs WHERE run = ?''', (self.run,))
        self.line = cur.fetchall()[0][0]
        cur.execute('''SELECT isoVar, lineVar, scaler, track FROM Runs WHERE run = ?''', (self.run,))
        self.var = cur.fetchall()[0]
        self.st = (ast.literal_eval(self.var[2]), ast.literal_eval(self.var[3]))  # tuple of (scaler and track)

        self.widthTG = 2
        self.centerTG = 1
        self.prepDB()

        # set files to fit
        cur.execute('''SELECT file FROM Files WHERE line = ? and type = ?''', (self.line, self.isotope,))
        paras = cur.fetchall()
        con.close()
        self.files = []
        for f in paras:
            self.files.append(f[0])

    def writeToFilerChi(self):
        ### write timeGates and rChi to excelfile

        # open Workbook to store result in
        try:
            wb = load_workbook(os.path.join(self.workingdir, 'FindTG.xlsx'))
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = 'rChi'
            ws['A1'] = 'TGcenter'
            ws['B1'] = 'TGwdith'
            ws['C1'] = 'rChi'
            ws['D1'] = 'centerCts'
            ws['E1'] = 'bgCts'
            wb.save(os.path.join(self.workingdir, 'FindTG.xlsx'))

        # get current TG from db
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT softwGateWidth FROM Runs WHERE run = ?''', (self.run,))
        width = cur.fetchall()[0][0]
        cur.execute('''SELECT midTof FROM Isotopes WHERE iso = ?''', (self.isotope,))
        midTOF = cur.fetchall()[0][0]

        # find results in database
        cur.execute('''SELECT rChi FROM FitRes WHERE iso = ? AND run = ?''', (self.isotope, self.run,))
        paras = cur.fetchall()
        print(paras)
        con.close()

        # calculate mean rChi
        sum = 0
        for x in paras:
            sum+=x[0]

        mean = sum/len(paras)
        print(mean)

        # write to Workbook
        max=row=ws.max_row
        ws.cell(row=max+1, column=1, value=midTOF)
        ws.cell(row=max+1, column=2, value=width)
        ws.cell(row=max+1, column=3, value=mean)

        wb.save(os.path.join(self.workingdir, 'FindTG.xlsx'))

    def writeToFile(self, snr):
        ### write timeGates and SNR to excelfile

        # open Workbook to store result in
        try:
            wb = load_workbook(os.path.join(self.workingdir, 'FindTG.xlsx'))
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = 'SNR'
            ws['A1'] = 'TGcenter'
            ws['B1'] = 'TGwdith'
            ws['C1'] = 'SNR'
            ws['D1'] = 'centerCts'
            ws['E1'] = 'bgCts'
            wb.save(os.path.join(self.workingdir, 'FindTG.xlsx'))

        # get current TG from db
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''SELECT softwGateWidth FROM Runs WHERE run = ?''', (self.run,))
        width = cur.fetchall()[0][0]
        cur.execute('''SELECT midTof FROM Isotopes WHERE iso = ?''', (self.isotope,))
        midTOF = cur.fetchall()[0][0]

        # write to Workbook
        max=row=ws.max_row
        ws.cell(row=max+1, column=1, value=midTOF)
        ws.cell(row=max+1, column=2, value=width)
        ws.cell(row=max+1, column=3, value=snr)

        wb.save(os.path.join(self.workingdir, 'FindTG.xlsx'))

    def prepDB(self):
        ### write TGcenter and TGwidth to Isotope and Runs
        con = sqlite3.connect(self.db)
        cur = con.cursor()
        cur.execute('''UPDATE Isotopes SET midTof = ? WHERE iso = ? ''', (self.centerTG, self.isotope,))
        cur.execute('''UPDATE Runs SET softwGateWidth = ? WHERE run = ?''', (self.widthTG, self.run,))
        con.commit()
        con.close()

    def shiftTG(self):
        ### shift TG by  µs
        self.centerTG += 0.5
        self.prepDB()

    def iterateCenter1(self):
        BatchFit.batchFit(self.files, self.db, self.run)
        self.writeToFile()
        while self.centerTG + self.widthTG/2 <= 10:
            self.shiftTG()
            BatchFit.batchFit(self.files, self.db, self.run)
            self.writeToFile()
        self.centerTG = 1

    def meanSNR(self):
        ### calculates the mean signal-to-noise ratio of all files for the current time gate

        snr = []  # list of signal-to-noise ratios

        con = sqlite3.connect(self.db)
        cur = con.cursor()

        ### loop through all files in self.files, fit and calculate noise
        for file in self.files:
            print('Fitting file:', file)
            cur.execute('''SELECT filePath FROM Files WHERE file = ?''', (file,))  # get path of file
            try:
                path = os.path.join(self.workingdir, cur.fetchall()[0][0])
            except:
                raise Exception(str(file) + 'not found in database.')

            ### load measurement
            softw_gates_trs = (self.db, self.run)  # define software gates for loaded data
            meas = MeasLoad.load(path, self.db, x_as_voltage=True, softw_gates=softw_gates_trs)

            ### create spectrum and fit
            iso = DBIsotope(self.db, meas.type, lineVar=self.var[1])    # get isotope from db
            spec = FullSpec(iso)    # create spectrum
            fit = SPFitter(spec, meas, self.st) # create object for fitting
            try:
                fit.fit()   # fit
            except RuntimeError:
                print('No fitparameters found. set SNR to 0')
                snr.append(0)
            else:
                ### calculate signal-to-noise ratio
                res = fit.calcRes() # calc residuals
                noise = np.std(res) # calc noise
                snr.append(fit.par[11] / noise) # append snr to list

        return  sum(snr) / len(snr) # return mean snr

    def iterateCenter(self):

        meanSNR = self.meanSNR()
        self.writeToFile(meanSNR)
        while self.centerTG + self.widthTG/2 <= 5:
            self.shiftTG()
            meanSNR = self.meanSNR()
            self.writeToFile(meanSNR)
        self.centerTG = 1

    def widenTG(self):
        ### make TG wider by 1
        self.widthTG += 0.5
        self.centerTG = self.widthTG / 2
        self.prepDB()

    def iterateWidth(self):
        self.iterateCenter()
        while self.centerTG + self.widthTG/2 <= 5:
            self.widenTG()
            self.iterateCenter()

    def findBestrChi(self):
    ### find best TG in excel-File

        # open Workbook to with results
        try:
            wb = load_workbook(os.path.join(self.workingdir, 'FindTG.xlsx'))
            ws = wb.active
        except:
            print('File not Found')
            return
        bestTG = (ws.cell('A2').value, ws.cell('B2').value)
        bestRChi = ws.cell('C2').value
        for row in ws.rows:
            print(row)
            print(row[2].value)
            if isinstance(row[2].value, str):
                pass
            elif abs(row[2].value -1) < abs(bestRChi-1):
                bestRChi = row[2].value
                bestTG = (row[0].value, row[1].value)
        print('Best TG is: ', bestTG)

    def findBest(self):
    ### find best TG in excel-File and set to db

        # open Workbook to with results
        try:
            wb = load_workbook(os.path.join(self.workingdir, 'FindTG.xlsx'))
            ws = wb.active
        except:
            print('File not Found')
            return
        bestTG = (ws.cell('A2').value, ws.cell('B2').value)
        bestSNR = ws.cell('C2').value
        for row in ws.rows:
            print(row)
            print(row[2].value)
            if isinstance(row[2].value, str):
                pass
            elif (abs(row[2].value) > abs(bestSNR) and abs(row[2].value) < 10):
                bestSNR = row[2].value
                bestTG = (row[0].value, row[1].value)
        print('Best TG is: ', bestTG, 'with SNR:', bestSNR)
        self.centerTG = bestTG[0]
        self.widthTG = bestTG[1]
        self.prepDB()

search = FindTG()
search.iterateWidth()
search.findBest()

